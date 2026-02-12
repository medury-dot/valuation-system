"""
Excel Report Generator for Valuation System (v3)
Generates a comprehensive .xlsx with:
- Assumptions as Single Source of Truth (all other sheets reference it)
- Cross-sheet formulas (no hardcoded values in DCF/Relative/Summary)
- Data Validation dropdowns (Sector Outlook, Peer Y/N)
- Ratio numerator/denominator in remarks
- 4 driver tabs (Macro 15%, Group 20%, Subgroup 35%, Company 30%)
- 10 sheets total (4-level driver hierarchy)
"""

import os
import re
import logging
import zipfile
from datetime import date, datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.formula import ArrayFormula
from dotenv import load_dotenv

load_dotenv(os.path.join(os.path.dirname(__file__), '..', 'config', '.env'))

logger = logging.getLogger(__name__)

# ── Styles ──────────────────────────────────────────────────────────────────
HEADER_FONT = Font(bold=True, size=12, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
SECTION_FONT = Font(bold=True, size=11, color='2F5496')
SECTION_FILL = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
INPUT_FILL = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
FORMULA_FILL = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
REMARK_FONT = Font(italic=True, color='808080', size=9)
ACTUAL_FONT = Font(color='006100', size=9)
DERIVED_FONT = Font(color='9C6500', size=9)
BOLD_FONT = Font(bold=True)
TITLE_FONT = Font(bold=True, size=14, color='2F5496')
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
# Log sheet styles
WARNING_FILL = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
ERROR_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
DEBUG_FONT = Font(color='808080', size=9)
LOG_FONT = Font(name='Courier New', size=9)

PCT_FMT = '0.00%'
NUM_FMT = '#,##0.00'
INT_FMT = '#,##0'


def _style_header(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
        # Set empty string to avoid invalid XML with styled but empty cells
        if cell.value is None:
            cell.value = ''


def _style_section(ws, row, max_col, label=None):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = SECTION_FONT
        cell.fill = SECTION_FILL
        # Set empty string to avoid invalid XML with styled but empty cells
        if cell.value is None:
            cell.value = ''
    if label:
        ws.cell(row=row, column=1, value=label)


def _inp(ws, row, col, value, fmt=None):
    """Write an input cell (yellow, editable)."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = INPUT_FILL
    cell.border = THIN_BORDER
    if fmt:
        cell.number_format = fmt
    return cell


def _fml(ws, row, col, formula, fmt=None):
    """Write a formula cell (green, computed)."""
    cell = ws.cell(row=row, column=col)
    cell.value = formula
    cell.fill = FORMULA_FILL
    cell.border = THIN_BORDER
    if fmt:
        cell.number_format = fmt
    return cell


def _remark(ws, row, col, text):
    """Write a remark/explanation."""
    # Prefix with single quote to prevent formula interpretation
    safe_text = f"'{text}" if text and (text.startswith('=') or text.startswith('+') or text.startswith('-')) else text
    cell = ws.cell(row=row, column=col, value=safe_text)
    cell.font = REMARK_FONT
    return cell


# Data source tag styles
ACTUAL_SOURCE_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # Green
DERIVED_SOURCE_FILL = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')  # Yellow
DEFAULT_SOURCE_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')  # Red


def _data_source_cell(ws, row, col, source_tag):
    """Write a data source tag cell with color coding."""
    if not source_tag:
        return
    cell = ws.cell(row=row, column=col, value=source_tag)
    cell.font = Font(bold=True, size=9)
    cell.alignment = Alignment(horizontal='center')
    cell.border = THIN_BORDER
    # Color code: green for actual, yellow for derived, red for default
    if 'ACTUAL' in source_tag:
        cell.fill = ACTUAL_SOURCE_FILL
        cell.font = Font(bold=True, size=9, color='006100')
    elif 'DERIVED' in source_tag:
        cell.fill = DERIVED_SOURCE_FILL
        cell.font = Font(bold=True, size=9, color='9C6500')
    elif 'DEFAULT' in source_tag:
        cell.fill = DEFAULT_SOURCE_FILL
        cell.font = Font(bold=True, size=9, color='9C0006')
    return cell


def _format_ratio_components(components, metric_label=''):
    """Build a readable remark string like 'avg(949/18148=5.2%, 850/16500=5.2%)'."""
    if not components:
        return ''
    parts = []
    for c in components:
        r = c.get('ratio', 0)
        n = c.get('numerator', c.get('nwc', 0))
        d = c.get('denominator', 0)
        parts.append(f"{n:,.0f}/{d:,.0f}={r:.1%}")
    avg_val = sum(c.get('ratio', 0) for c in components) / len(components)
    return f"avg({', '.join(parts)}) = {avg_val:.1%}"


def _format_nwc_components(components):
    """Build NWC remark like 'avg((980+1200-2800)/18148=-3.4%)'."""
    if not components:
        return ''
    parts = []
    for c in components:
        parts.append(
            f"({c['inv']:,.0f}+{c['debtors']:,.0f}-{c['payables']:,.0f})/"
            f"{c['denominator']:,.0f}={c['ratio']:.1%}"
        )
    avg_val = sum(c['ratio'] for c in components) / len(components)
    return f"avg({', '.join(parts)}) = {avg_val:.1%}"


def _format_tax_components(components):
    """Build tax remark like 'avg(1-3800/5100=25.5%)'."""
    if not components:
        return ''
    parts = []
    for c in components:
        parts.append(f"1-{c['pat']:,.0f}/{c['pbt']:,.0f}={c['rate']:.1%}")
    avg_val = sum(c['rate'] for c in components) / len(components)
    return f"avg({', '.join(parts)}) = {avg_val:.1%}"


def _format_cod_components(components):
    """Build cost of debt remark like 'avg(120/800=15.0%)'."""
    if not components:
        return ''
    parts = []
    for c in components:
        parts.append(f"{c['interest']:,.0f}/{c['debt']:,.0f}={c['rate']:.1%}")
    avg_val = sum(c['rate'] for c in components) / len(components)
    return f"avg({', '.join(parts)}) = {avg_val:.1%}"


# ── Main Entry Point ────────────────────────────────────────────────────────

def generate_valuation_excel(result: dict, output_path: str = None) -> str:
    """
    Generate a comprehensive Excel valuation report.

    Args:
        result: Output from ValuatorAgent.run_full_valuation()
        output_path: Optional output path. Defaults to logs/ with timestamp.

    Returns:
        Path to generated Excel file.
    """
    company = result.get('company_name', 'Unknown')
    symbol = result.get('nse_symbol', '')

    if not output_path:
        log_dir = os.path.join(os.path.dirname(__file__), '..', 'logs')
        os.makedirs(log_dir, exist_ok=True)
        safe_name = symbol.lower() if symbol else company.lower().replace(' ', '_')
        timestamp = datetime.now().strftime('%Y%m%d_%H%M')
        output_path = os.path.join(log_dir, f'{safe_name}_valuation_{timestamp}.xlsx')

    wb = Workbook()
    # Cell registry: tracks key cells for cross-sheet references
    refs = {}

    # Build all sheets in order
    ws_assumptions = wb.create_sheet('Assumptions')
    _build_assumptions_sheet(ws_assumptions, result, refs)

    # WEEK 1 FIX: Add Beta Scenarios sheet
    ws_beta = wb.create_sheet('Beta Scenarios')
    _build_beta_scenarios_sheet(ws_beta, result, refs)

    ws_dcf = wb.create_sheet('DCF Model')
    _build_dcf_sheet(ws_dcf, result, refs)

    ws_rel = wb.create_sheet('Relative Val')
    _build_relative_sheet(ws_rel, result, refs)

    # Patch Assumptions peer medians to reference Relative Val filtered medians
    if refs.get('filtered_median_row'):
        fmr = refs['filtered_median_row']
        rel_sheet = "'Relative Val'"
        # I=PE, J=PB, K=EV/EBITDA, L=P/S in the peer table
        for ref_key, col_letter in [('peer_pe_median', 'I'), ('peer_pb_median', 'J'),
                                     ('peer_eveb_median', 'K'), ('peer_ps_median', 'L')]:
            assumptions_row = refs.get(ref_key)
            if assumptions_row:
                cell = ws_assumptions.cell(row=assumptions_row, column=2)
                cell.value = f'={rel_sheet}!{col_letter}{fmr}'
                cell.number_format = '0.00'

    ws_mc = wb.create_sheet('Monte Carlo')
    _build_monte_carlo_sheet(ws_mc, result, refs)

    ws_sens = wb.create_sheet('Sensitivity')
    _build_sensitivity_sheet(ws_sens, result, refs)

    # Consolidated Driver Hierarchy sheet (replaces 4 separate driver sheets)
    ws_drivers = wb.create_sheet('Driver Hierarchy')
    _build_driver_hierarchy_sheet(ws_drivers, result, refs)

    # Computation Log sheet (captures all log entries from the valuation run)
    ws_log = wb.create_sheet('Computation Log')
    _build_log_sheet(ws_log, result, refs)

    # Summary is the active (first) sheet — built last since it references others
    ws_summary = wb.active
    ws_summary.title = 'Summary'
    _build_summary_sheet(ws_summary, result, refs)

    # Move Summary to position 0 (now 9 sheets: Summary + 8 others)
    wb.move_sheet('Summary', offset=-8)

    wb.save(output_path)

    # Post-process: Fix openpyxl bug where formula cells get empty <v></v> tags
    # which can cause Excel repair warnings
    _fix_empty_formula_values(output_path)

    logger.info(f"Valuation Excel saved to: {output_path}")
    return output_path


def _fix_empty_formula_values(xlsx_path: str):
    """
    Remove empty <v></v> tags from formula cells in xlsx.
    openpyxl generates these but some Excel versions don't like them.
    """
    temp_path = xlsx_path + '.tmp'
    try:
        with zipfile.ZipFile(xlsx_path, 'r') as zin:
            with zipfile.ZipFile(temp_path, 'w', zipfile.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    data = zin.read(item.filename)
                    if item.filename.startswith('xl/worksheets/') and item.filename.endswith('.xml'):
                        content = data.decode('utf-8')
                        # Remove empty <v></v> tags after formula elements
                        content = re.sub(r'(<f>[^<]*</f>)<v></v>', r'\1', content)
                        data = content.encode('utf-8')
                    zout.writestr(item, data)
        os.replace(temp_path, xlsx_path)
    except Exception as e:
        logger.warning(f"Failed to fix formula values in xlsx: {e}")
        if os.path.exists(temp_path):
            os.remove(temp_path)


# ── Sheet 2: Assumptions (Single Source of Truth) ───────────────────────────

def _build_assumptions_sheet(ws, result, refs):
    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 60
    ws.column_dimensions['D'].width = 18  # Data Source column

    assumptions = result.get('dcf_assumptions', {})
    dcf_details = result.get('dcf_details', {})
    dcf_assum = dcf_details.get('assumptions', {})
    data_sources = result.get('data_sources', {})
    outlook = result.get('sector_outlook', {})

    r = 1
    ws.cell(row=r, column=1, value='Valuation Assumptions').font = TITLE_FONT
    _remark(ws, r, 3, 'Yellow = changeable input, Green = formula/computed')

    # ── MACRO LEVEL ──
    r = 3
    _style_section(ws, r, 3, 'MACRO LEVEL')
    _remark(ws, r, 3, 'Weight: 15% of driver hierarchy')
    r += 1
    ws.cell(row=r, column=1, value='Risk-Free Rate (Rf)')
    refs['rf_rate'] = r
    _inp(ws, r, 2, assumptions.get('risk_free_rate', dcf_assum.get('risk_free_rate', 0.0674)), PCT_FMT)
    _remark(ws, r, 3, '[ACTUAL: interest_rate_10y from market_indicators.csv]')
    r += 1
    ws.cell(row=r, column=1, value='Equity Risk Premium (ERP)')
    refs['erp'] = r
    _inp(ws, r, 2, assumptions.get('equity_risk_premium', assumptions.get('erp', dcf_assum.get('erp', 0.0708))), PCT_FMT)
    _remark(ws, r, 3, '[ACTUAL: India equity risk premium]')

    # ── GROUP LEVEL ──
    r += 3
    _style_section(ws, r, 3, 'GROUP LEVEL')
    _remark(ws, r, 3, 'Weight: 20% of driver hierarchy (valuation_group)')
    r += 1
    ws.cell(row=r, column=1, value='Beta (Levered)')
    refs['beta'] = r
    _inp(ws, r, 2, assumptions.get('beta', dcf_assum.get('beta', 0)), '0.0000')
    _remark(ws, r, 3, '[ACTUAL: Sector beta]')

    r += 1
    ws.cell(row=r, column=1, value='Sector Outlook')
    refs['outlook_label'] = r
    outlook_label = outlook.get('outlook_label', outlook.get('outlook', 'NEUTRAL'))
    outlook_cell = _inp(ws, r, 2, outlook_label)
    # Data Validation dropdown
    dv = DataValidation(type='list', formula1='"BEARISH,NEGATIVE,NEUTRAL,POSITIVE,BULLISH"',
                        allow_blank=False)
    dv.error = 'Select from: BEARISH, NEGATIVE, NEUTRAL, POSITIVE, BULLISH'
    dv.errorTitle = 'Invalid Outlook'
    ws.add_data_validation(dv)
    dv.add(outlook_cell)
    _remark(ws, r, 3, 'Dropdown: change to recalculate growth/margin adjustments')

    # Lookup table for outlook → score (write first before referencing)
    lookup_data = [('BEARISH', -1.0), ('NEGATIVE', -0.5), ('NEUTRAL', 0.0),
                   ('POSITIVE', 0.5), ('BULLISH', 1.0)]
    for i, (label, score) in enumerate(lookup_data):
        ws.cell(row=80 + i, column=5, value=label)
        ws.cell(row=80 + i, column=6, value=score)
    refs['lookup_range'] = 'E80:F84'

    r += 1
    ws.cell(row=r, column=1, value='Outlook Score')
    refs['outlook_score'] = r
    # Use nested IFs instead of VLOOKUP to avoid formula errors
    outlook_ref = refs["outlook_label"]
    nested_if = f'=IF(B{outlook_ref}="BULLISH",1,IF(B{outlook_ref}="POSITIVE",0.5,IF(B{outlook_ref}="NEUTRAL",0,IF(B{outlook_ref}="NEGATIVE",-0.5,-1))))'
    _fml(ws, r, 2, nested_if, '0.00')
    _remark(ws, r, 3, 'Maps label to score: BEARISH=-1.0 ... BULLISH=+1.0')

    r += 1
    ws.cell(row=r, column=1, value='Growth Adjustment')
    refs['growth_adj'] = r
    _fml(ws, r, 2, f'=B{refs["outlook_score"]}*0.03', PCT_FMT)
    _remark(ws, r, 3, '= Outlook Score * 3% (max +/-3%)')

    r += 1
    ws.cell(row=r, column=1, value='Margin Adjustment')
    refs['margin_adj'] = r
    _fml(ws, r, 2, f'=B{refs["outlook_score"]}*0.01', PCT_FMT)
    _remark(ws, r, 3, '= Outlook Score * 1% (max +/-1%)')

    # Peer medians (will be ArrayFormulas referencing Relative Val Y/N)
    r += 1
    ws.cell(row=r, column=1, value='Peer PE Median')
    refs['peer_pe_median'] = r
    # Placeholder — will be set after Relative Val sheet builds
    _inp(ws, r, 2, _get_peer_median(result, 'pe'), '0.00')
    _remark(ws, r, 3, 'From included peers (Y/N toggle in Relative Val)')

    r += 1
    ws.cell(row=r, column=1, value='Peer PB Median')
    refs['peer_pb_median'] = r
    _inp(ws, r, 2, _get_peer_median(result, 'pb'), '0.00')

    r += 1
    ws.cell(row=r, column=1, value='Peer EV/EBITDA Median')
    refs['peer_eveb_median'] = r
    _inp(ws, r, 2, _get_peer_median(result, 'ev_ebitda'), '0.00')

    r += 1
    ws.cell(row=r, column=1, value='Peer PS Median')
    refs['peer_ps_median'] = r
    _inp(ws, r, 2, _get_peer_median(result, 'ps'), '0.00')

    # ── SUBGROUP LEVEL ──
    r += 3
    _style_section(ws, r, 3, 'SUBGROUP LEVEL')
    _remark(ws, r, 3, 'Weight: 35% of driver hierarchy (valuation_subgroup)')

    r += 1
    ws.cell(row=r, column=1, value='Valuation Subgroup')
    refs['valuation_subgroup'] = r
    subgroup = result.get('valuation_subgroup', outlook.get('valuation_subgroup', ''))
    _inp(ws, r, 2, subgroup or '(none)')
    _remark(ws, r, 3, 'E.g., INDUSTRIALS_DEFENSE, INDUSTRIALS_ENGINEERING')

    r += 1
    ws.cell(row=r, column=1, value='Subgroup Score')
    refs['subgroup_score'] = r
    subgroup_score = result.get('subgroup_outlook', outlook.get('subgroup_score', 0))
    _inp(ws, r, 2, subgroup_score, '0.0000')
    _remark(ws, r, 3, 'Weighted avg of subgroup-specific drivers')

    r += 1
    ws.cell(row=r, column=1, value='Subgroup Growth Adjustment')
    refs['subgroup_growth_adj'] = r
    _fml(ws, r, 2, f'=B{refs["subgroup_score"]}*0.03', PCT_FMT)
    _remark(ws, r, 3, '= Subgroup Score * 3% (max +/-3%)')

    r += 1
    ws.cell(row=r, column=1, value='Subgroup Margin Adjustment')
    refs['subgroup_margin_adj'] = r
    _fml(ws, r, 2, f'=B{refs["subgroup_score"]}*0.01', PCT_FMT)
    _remark(ws, r, 3, '= Subgroup Score * 1% (max +/-1%)')

    r += 1
    ws.cell(row=r, column=1, value='Combined Group+Subgroup Score')
    refs['combined_score'] = r
    combined = outlook.get('outlook_score', 0)
    _fml(ws, r, 2, f'=(B{refs["outlook_score"]}*0.20 + B{refs["subgroup_score"]}*0.35)/0.55', '0.0000')
    _remark(ws, r, 3, '(Group×20% + Subgroup×35%) / 55%')

    r += 1
    ws.cell(row=r, column=1, value='Total Growth Adjustment')
    refs['total_growth_adj'] = r
    _fml(ws, r, 2, f'=B{refs["growth_adj"]}*0.20/0.55 + B{refs["subgroup_growth_adj"]}*0.35/0.55', PCT_FMT)
    _remark(ws, r, 3, 'Weighted: Group 20% + Subgroup 35%')

    r += 1
    ws.cell(row=r, column=1, value='Total Margin Adjustment')
    refs['total_margin_adj'] = r
    _fml(ws, r, 2, f'=B{refs["margin_adj"]}*0.20/0.55 + B{refs["subgroup_margin_adj"]}*0.35/0.55', PCT_FMT)
    _remark(ws, r, 3, 'Weighted: Group 20% + Subgroup 35%')

    # ── COMPANY LEVEL ──
    r += 3
    _style_section(ws, r, 3, 'COMPANY LEVEL')
    _remark(ws, r, 3, 'Weight: 30% of driver hierarchy')

    r += 1
    ws.cell(row=r, column=1, value='Revenue Base (Rs Cr)')
    refs['base_revenue'] = r
    rev = assumptions.get('base_revenue', dcf_assum.get('base_revenue', 0))
    _inp(ws, r, 2, rev, NUM_FMT)
    _remark(ws, r, 3, f'[{result.get("revenue_source", "")}]')

    # CAGR
    r += 1
    ws.cell(row=r, column=1, value='Revenue CAGR 3Y')
    refs['cagr_3y'] = r
    _inp(ws, r, 2, result.get('revenue_cagr_3y') or 0, PCT_FMT)

    r += 1
    ws.cell(row=r, column=1, value='Revenue CAGR 5Y')
    refs['cagr_5y'] = r
    _inp(ws, r, 2, result.get('revenue_cagr_5y') or 0, PCT_FMT)

    # YoY growth
    yoy = result.get('revenue_yoy_growth', [])
    for item in yoy:
        r += 1
        ws.cell(row=r, column=1, value=f'YoY Growth FY{item["to_year"]}')
        _inp(ws, r, 2, item['growth'], PCT_FMT)
        _remark(ws, r, 3,
                f'= ({item["to_value"]:,.0f} / {item["from_value"]:,.0f}) - 1 = {item["growth"]:.1%}')

    # Growth trajectory with formulas
    r += 1
    ws.cell(row=r, column=1, value='Growth Y1 (base)')
    refs['growth_y1_base'] = r
    _fml(ws, r, 2, f'=B{refs["cagr_3y"]}*0.6+B{refs["cagr_5y"]}*0.4', PCT_FMT)
    _remark(ws, r, 3, '= CAGR_3Y * 0.6 + CAGR_5Y * 0.4')

    r += 1
    ws.cell(row=r, column=1, value='Growth Y1 (sector adj)')
    refs['growth_y1'] = r
    _fml(ws, r, 2, f'=MAX(0.03,B{refs["growth_y1_base"]}*(1+B{refs["growth_adj"]}))', PCT_FMT)
    _remark(ws, r, 3, '= MAX(3%, base * (1 + sector growth adj))')

    for yr in range(2, 6):
        r += 1
        ws.cell(row=r, column=1, value=f'Growth Y{yr} (sector adj)')
        refs[f'growth_y{yr}'] = r
        _fml(ws, r, 2,
             f'=MAX(0.03,B{refs["growth_y1"]}-(B{refs["growth_y1"]}-0.06)*({yr-1}/4))', PCT_FMT)
        _remark(ws, r, 3, f'= MAX(3%, Y1 - (Y1-6%) * {yr-1}/4) — linear decay to 6%')

    # TTM PAT / PBIDT
    r += 1
    ws.cell(row=r, column=1, value='TTM PAT (Rs Cr)')
    refs['ttm_pat'] = r
    _inp(ws, r, 2, result.get('ttm_pat', 0), NUM_FMT)
    _remark(ws, r, 3, f'[{result.get("pat_source", "")}]')

    r += 1
    ws.cell(row=r, column=1, value='TTM PBIDT (Rs Cr)')
    refs['ttm_pbidt'] = r
    _inp(ws, r, 2, result.get('ttm_pbidt', 0), NUM_FMT)
    _remark(ws, r, 3, f'[{result.get("pbidt_source", "")}]')

    # EBITDA Margin as formula
    r += 1
    ws.cell(row=r, column=1, value='EBITDA Margin')
    refs['ebitda_margin'] = r
    ebitda_m = assumptions.get('ebitda_margin', dcf_assum.get('ebitda_margin', 0.15))
    _inp(ws, r, 2, ebitda_m, PCT_FMT)
    _remark(ws, r, 3, f'= TTM PBIDT / TTM Revenue (or avg 3yr)')

    r += 1
    ws.cell(row=r, column=1, value='Margin Improvement (annual)')
    refs['margin_improvement'] = r
    _inp(ws, r, 2, assumptions.get('margin_improvement', dcf_assum.get('margin_improvement', 0)), '0.0000')

    # Ratios with numerator/denominator in remarks
    r += 1
    ws.cell(row=r, column=1, value='Capex / Sales')
    refs['capex_to_sales'] = r
    capex_s = assumptions.get('capex_to_sales', dcf_assum.get('capex_to_sales', 0.08))
    _inp(ws, r, 2, capex_s, PCT_FMT)
    _remark(ws, r, 3, _format_ratio_components(result.get('capex_components', [])))

    r += 1
    ws.cell(row=r, column=1, value='Depreciation / Sales')
    refs['depr_to_sales'] = r
    depr_s = assumptions.get('depreciation_to_sales', dcf_assum.get('depreciation_to_sales', 0.04))
    _inp(ws, r, 2, depr_s, PCT_FMT)
    _remark(ws, r, 3, _format_ratio_components(result.get('depr_components', [])))

    r += 1
    ws.cell(row=r, column=1, value='NWC / Sales')
    refs['nwc_to_sales'] = r
    nwc_s = assumptions.get('nwc_to_sales', dcf_assum.get('nwc_to_sales', 0.15))
    _inp(ws, r, 2, nwc_s, PCT_FMT)
    _remark(ws, r, 3, _format_nwc_components(result.get('nwc_components', [])))
    _data_source_cell(ws, r, 4, data_sources.get('nwc', ''))

    r += 1
    ws.cell(row=r, column=1, value='Effective Tax Rate')
    refs['tax_rate'] = r
    _inp(ws, r, 2, assumptions.get('tax_rate', dcf_assum.get('tax_rate', 0.25)), PCT_FMT)
    _remark(ws, r, 3, _format_tax_components(result.get('tax_components', [])))
    _data_source_cell(ws, r, 4, data_sources.get('tax', ''))

    r += 1
    ws.cell(row=r, column=1, value='Cost of Debt (pre-tax)')
    refs['cost_of_debt'] = r
    _inp(ws, r, 2, assumptions.get('cost_of_debt', dcf_assum.get('cost_of_debt', 0.09)), PCT_FMT)
    _remark(ws, r, 3, _format_cod_components(result.get('cost_of_debt_components', [])))

    r += 1
    ws.cell(row=r, column=1, value='Debt Ratio (D/V)')
    refs['debt_ratio'] = r
    _inp(ws, r, 2, assumptions.get('debt_ratio', dcf_assum.get('debt_ratio', 0.10)), PCT_FMT)

    r += 1
    ws.cell(row=r, column=1, value='Terminal ROCE')
    refs['terminal_roce'] = r
    _inp(ws, r, 2, assumptions.get('terminal_roce', dcf_assum.get('terminal_roce', 0.15)), PCT_FMT)
    _data_source_cell(ws, r, 4, data_sources.get('roce', ''))

    r += 1
    ws.cell(row=r, column=1, value='Terminal Reinvestment Rate')
    refs['terminal_reinvestment'] = r
    _inp(ws, r, 2, assumptions.get('terminal_reinvestment', dcf_assum.get('terminal_reinvestment', 0.30)), PCT_FMT)
    _data_source_cell(ws, r, 4, data_sources.get('reinvest', ''))

    r += 1
    ws.cell(row=r, column=1, value='Net Debt (Rs Cr)')
    refs['net_debt'] = r
    _inp(ws, r, 2, assumptions.get('net_debt', dcf_assum.get('net_debt', 0)), NUM_FMT)

    r += 1
    ws.cell(row=r, column=1, value='Cash & Equivalents (Rs Cr)')
    refs['cash'] = r
    cash_val = assumptions.get('cash_and_equivalents', dcf_assum.get('cash_and_equivalents', 0))
    _inp(ws, r, 2, cash_val, NUM_FMT)
    _remark(ws, r, 3, f'[ACTUAL: {result.get("cash_source_detail", "")}]')

    r += 1
    ws.cell(row=r, column=1, value='Shares Outstanding (Cr)')
    refs['shares'] = r
    _inp(ws, r, 2, assumptions.get('shares_outstanding', dcf_assum.get('shares_outstanding', 0)), '0.0000')
    shares_remark = '= MCap / CMP' if data_sources.get('shares') != 'ACTUAL_COLUMN' else '= Actual paid-up shares'
    _remark(ws, r, 3, shares_remark)
    _data_source_cell(ws, r, 4, data_sources.get('shares', ''))


def _get_peer_median(result, key):
    """Extract peer median from result dict."""
    rel = result.get('relative_details', {})
    implied = rel.get('implied_values', {})
    return implied.get(key, {}).get('peer_median', 0) or 0


# ── Sheet 3: Beta Scenarios ────────────────────────────────────────────────

def _build_beta_scenarios_sheet(ws, result, refs):
    """
    Week 1 FIX: Beta scenario comparison sheet.
    Shows 3 beta scenarios (A/B/C) side-by-side with:
    - Beta calculations (unlevered and levered)
    - WACC breakdown
    - DCF intrinsic values
    - Comparison to CMP and BASE DCF
    """
    ws.column_dimensions['A'].width = 28
    for c in 'BCDEFGH':
        ws.column_dimensions[c].width = 18

    # Get beta scenarios from result
    beta_scenarios = result.get('dcf_beta_scenarios', {})

    # WEEK 1 DEBUG: Log what's in result
    logger.info(f"Beta Scenarios sheet: dcf_beta_scenarios key exists = {'dcf_beta_scenarios' in result}")
    logger.info(f"Beta Scenarios sheet: number of scenarios = {len(beta_scenarios)}")
    if beta_scenarios:
        logger.info(f"Beta Scenarios sheet: scenario keys = {list(beta_scenarios.keys())}")

    if not beta_scenarios:
        # No beta scenarios available
        ws.cell(row=1, column=1, value='Beta Scenarios Not Available').font = TITLE_FONT
        ws.cell(row=2, column=1, value='Beta scenarios are only computed in full valuation mode.')
        ws.cell(row=3, column=1, value=f'Debug: Result keys = {list(result.keys())[:20]}').font = REMARK_FONT
        return

    # Title
    company = result.get('company_name', 'Unknown')
    symbol = result.get('nse_symbol', '')
    ws.cell(row=1, column=1, value=f'Beta Scenario Analysis: {company}').font = TITLE_FONT
    ws.merge_cells('A1:H1')

    # Company info
    r = 3
    ws.cell(row=r, column=1, value='Symbol:').font = BOLD_FONT
    ws.cell(row=r, column=2, value=symbol)
    ws.cell(row=r, column=3, value='CMP:').font = BOLD_FONT
    cmp = result.get('cmp', 0)
    ws.cell(row=r, column=4, value=cmp).number_format = NUM_FMT
    ws.cell(row=r, column=5, value='BASE DCF:').font = BOLD_FONT
    base_dcf = result.get('dcf_base', 0)
    ws.cell(row=r, column=6, value=base_dcf).number_format = NUM_FMT

    r += 1
    ws.cell(row=r, column=1, value='Valuation Group:').font = BOLD_FONT
    ws.cell(row=r, column=2, value=result.get('valuation_group', ''))
    ws.cell(row=r, column=3, value='Valuation Subgroup:').font = BOLD_FONT
    ws.cell(row=r, column=4, value=result.get('valuation_subgroup', ''))
    ws.merge_cells(f'D{r}:F{r}')

    # Section: Beta Scenario Comparison
    r += 2
    _style_section(ws, r, 8, 'BETA SCENARIO COMPARISON')

    r += 1
    # Headers
    headers = ['Scenario', 'Beta (Levered)', 'Beta (Unlevered)', 'WACC', 'Cost of Equity', 'DCF Intrinsic', 'vs CMP', 'Beta Source']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=r, column=col_idx, value=header)
        cell.font = BOLD_FONT
        cell.fill = SECTION_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')

    # Scenario rows
    scenario_labels = {
        'individual_weekly': 'A: Individual (Weekly)',
        'damodaran_india': 'B: Damodaran (India)',
        'subgroup_aggregate': 'C: Subgroup (Aggregate)'
    }

    scenario_colors = {
        'individual_weekly': PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid'),  # Green
        'damodaran_india': PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid'),    # Yellow
        'subgroup_aggregate': PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')  # Blue
    }

    for scenario_key in ['individual_weekly', 'damodaran_india', 'subgroup_aggregate']:
        if scenario_key not in beta_scenarios:
            continue

        r += 1
        scenario_data = beta_scenarios[scenario_key]
        label = scenario_labels.get(scenario_key, scenario_key)
        fill = scenario_colors.get(scenario_key, FORMULA_FILL)

        # Scenario name
        cell = ws.cell(row=r, column=1, value=label)
        cell.font = BOLD_FONT
        cell.fill = fill
        cell.border = THIN_BORDER

        # Beta (levered)
        cell = ws.cell(row=r, column=2, value=scenario_data.get('beta', 0))
        cell.number_format = '0.000'
        cell.fill = fill
        cell.border = THIN_BORDER

        # Beta (unlevered)
        cell = ws.cell(row=r, column=3, value=scenario_data.get('beta_unlevered', 0))
        cell.number_format = '0.000'
        cell.fill = fill
        cell.border = THIN_BORDER

        # WACC
        cell = ws.cell(row=r, column=4, value=scenario_data.get('wacc', 0))
        cell.number_format = PCT_FMT
        cell.fill = fill
        cell.border = THIN_BORDER

        # Cost of Equity
        cell = ws.cell(row=r, column=5, value=scenario_data.get('cost_of_equity', 0))
        cell.number_format = PCT_FMT
        cell.fill = fill
        cell.border = THIN_BORDER

        # DCF Intrinsic
        dcf_value = scenario_data.get('intrinsic_value', 0)
        cell = ws.cell(row=r, column=6, value=dcf_value)
        cell.number_format = NUM_FMT
        cell.fill = fill
        cell.border = THIN_BORDER

        # vs CMP
        gap_vs_cmp = ((dcf_value / cmp) - 1) if cmp > 0 else 0
        cell = ws.cell(row=r, column=7, value=gap_vs_cmp)
        cell.number_format = PCT_FMT
        cell.fill = fill
        cell.border = THIN_BORDER
        # Conditional formatting: green if positive, red if negative
        if gap_vs_cmp > 0:
            cell.font = Font(color='006100', bold=True)
        else:
            cell.font = Font(color='9C0006', bold=True)

        # Beta Source
        source = scenario_data.get('beta_source', '')
        # Format for display
        if scenario_key == 'damodaran_india':
            industry = scenario_data.get('industry', '')
            n_firms = scenario_data.get('n_firms', '')
            if industry:
                source = f'{industry} (India, {n_firms} firms)'
        cell = ws.cell(row=r, column=8, value=source)
        cell.fill = fill
        cell.border = THIN_BORDER

    # Add BASE DCF for comparison
    r += 1
    ws.cell(row=r, column=1, value='BASE (Current DCF)').font = Font(bold=True, color='FF0000')
    dcf_assumptions = result.get('dcf_assumptions', {})
    base_beta = dcf_assumptions.get('beta', 0)
    base_wacc = dcf_assumptions.get('wacc', 0)
    ws.cell(row=r, column=2, value=base_beta).number_format = '0.000'
    ws.cell(row=r, column=4, value=base_wacc).number_format = PCT_FMT
    ws.cell(row=r, column=6, value=base_dcf).number_format = NUM_FMT
    base_gap = ((base_dcf / cmp) - 1) if cmp > 0 else 0
    cell = ws.cell(row=r, column=7, value=base_gap)
    cell.number_format = PCT_FMT
    if base_gap > 0:
        cell.font = Font(color='006100', bold=True)
    else:
        cell.font = Font(color='9C0006', bold=True)

    # Section: Recommendation
    r += 2
    _style_section(ws, r, 8, 'RECOMMENDATION')

    r += 1
    # Find best scenario (closest to CMP or highest DCF if undervalued)
    best_scenario = None
    best_gap = -999999
    for key, data in beta_scenarios.items():
        dcf_val = data.get('intrinsic_value', 0)
        gap = ((dcf_val / cmp) - 1) if cmp > 0 else -999999
        if gap > best_gap:
            best_gap = gap
            best_scenario = key

    if best_scenario:
        label = scenario_labels.get(best_scenario, best_scenario)
        best_dcf = beta_scenarios[best_scenario].get('intrinsic_value', 0)
        improvement = ((best_dcf / base_dcf) - 1) * 100 if base_dcf > 0 else 0

        ws.cell(row=r, column=1, value='Recommended Scenario:').font = BOLD_FONT
        ws.cell(row=r, column=2, value=label).font = Font(bold=True, color='006100')
        ws.merge_cells(f'B{r}:D{r}')

        r += 1
        ws.cell(row=r, column=1, value='Improvement over BASE:')
        cell = ws.cell(row=r, column=2, value=improvement / 100)
        cell.number_format = PCT_FMT
        cell.font = Font(bold=True, color='006100' if improvement > 0 else '9C0006')

        r += 1
        ws.cell(row=r, column=1, value='Gap vs CMP:')
        cell = ws.cell(row=r, column=2, value=best_gap)
        cell.number_format = PCT_FMT
        cell.font = Font(bold=True, color='006100' if best_gap > 0 else '9C0006')

    # Section: Beta Methodology Explanation
    r += 2
    _style_section(ws, r, 8, 'BETA SCENARIO METHODOLOGY')

    r += 1
    methodology = [
        ('Scenario A (Individual)', 'Company-specific beta from 2yr+5yr NIFTY regression. Most accurate when available.'),
        ('Scenario B (Damodaran India)', 'India industry beta from Damodaran dataset (93 industries). Professional estimate based on NSE/BSE sample.'),
        ('Scenario C (Subgroup Aggregate)', 'Peer average beta from valuation subgroup (35-100 companies). Broadest comparison.'),
        ('', ''),
        ('Recommendation', 'Use Scenario A when available and company has >2yr history. Use Scenario B for newly listed or volatile companies. Use Scenario C only as fallback.'),
    ]

    for desc, explanation in methodology:
        ws.cell(row=r, column=1, value=desc).font = BOLD_FONT
        ws.cell(row=r, column=2, value=explanation)
        ws.merge_cells(f'B{r}:H{r}')
        r += 1

    # Section: WACC Breakdown (for reference)
    r += 1
    _style_section(ws, r, 8, 'WACC CALCULATION BREAKDOWN')

    r += 1
    ws.cell(row=r, column=1, value='Component').font = BOLD_FONT
    ws.cell(row=r, column=2, value='Scenario A').font = BOLD_FONT
    ws.cell(row=r, column=3, value='Scenario B').font = BOLD_FONT
    ws.cell(row=r, column=4, value='Scenario C').font = BOLD_FONT
    ws.cell(row=r, column=5, value='BASE').font = BOLD_FONT
    _style_header(ws, r, 5)

    # Get common parameters
    rf = dcf_assumptions.get('risk_free_rate', 0)
    erp = dcf_assumptions.get('erp', 0)
    kd = dcf_assumptions.get('cost_of_debt_at', 0)
    de_ratio = dcf_assumptions.get('debt_ratio', 0) / (1 - dcf_assumptions.get('debt_ratio', 0.2)) if dcf_assumptions.get('debt_ratio', 0) < 1 else 0.2
    tax_rate = dcf_assumptions.get('tax_rate', 0.25)

    # Build WACC breakdown rows
    breakdown_rows = [
        ('Risk-free Rate', rf, PCT_FMT),
        ('Equity Risk Premium', erp, PCT_FMT),
        ('Beta (Levered)', None, '0.000'),  # Will fill per scenario
        ('Cost of Equity (Ke)', None, PCT_FMT),  # Will calculate
        ('Cost of Debt (Kd)', kd, PCT_FMT),
        ('Tax Rate', tax_rate, PCT_FMT),
        ('D/E Ratio', de_ratio, '0.00'),
        ('Equity Weight', 1/(1+de_ratio), PCT_FMT),
        ('Debt Weight', de_ratio/(1+de_ratio), PCT_FMT),
        ('WACC', None, PCT_FMT),  # Will fill per scenario
    ]

    for label, base_value, fmt in breakdown_rows:
        r += 1
        ws.cell(row=r, column=1, value=label)

        # Fill values for each scenario
        for col_offset, scenario_key in enumerate(['individual_weekly', 'damodaran_india', 'subgroup_aggregate'], start=0):
            if scenario_key not in beta_scenarios:
                continue

            data = beta_scenarios[scenario_key]

            if label == 'Beta (Levered)':
                val = data.get('beta', 0)
            elif label == 'Cost of Equity (Ke)':
                val = data.get('cost_of_equity', 0)
            elif label == 'WACC':
                val = data.get('wacc', 0)
            else:
                val = base_value  # Common values

            if val is not None:
                ws.cell(row=r, column=2+col_offset, value=val).number_format = fmt

        # BASE column
        if label == 'Beta (Levered)':
            val = base_beta
        elif label == 'Cost of Equity (Ke)':
            val = base_wacc / ((1/(1+de_ratio)) + (de_ratio/(1+de_ratio)) * kd * (1-tax_rate))  # Back-calc Ke
        elif label == 'WACC':
            val = base_wacc
        else:
            val = base_value

        if val is not None:
            ws.cell(row=r, column=5, value=val).number_format = fmt

    # Section: Sensitivity Analysis
    r += 2
    _style_section(ws, r, 8, 'SENSITIVITY ANALYSIS')

    r += 1
    ws.cell(row=r, column=1, value='Impact of ±2% WACC change on DCF intrinsic value:')
    ws.merge_cells(f'A{r}:H{r}')

    r += 1
    ws.cell(row=r, column=1, value='Scenario').font = BOLD_FONT
    ws.cell(row=r, column=2, value='WACC -2%').font = BOLD_FONT
    ws.cell(row=r, column=3, value='WACC -1%').font = BOLD_FONT
    ws.cell(row=r, column=4, value='Base WACC').font = BOLD_FONT
    ws.cell(row=r, column=5, value='WACC +1%').font = BOLD_FONT
    ws.cell(row=r, column=6, value='WACC +2%').font = BOLD_FONT
    _style_header(ws, r, 6)

    for scenario_key in ['individual_weekly', 'damodaran_india', 'subgroup_aggregate']:
        if scenario_key not in beta_scenarios:
            continue

        r += 1
        data = beta_scenarios[scenario_key]
        label = scenario_labels.get(scenario_key, scenario_key)

        ws.cell(row=r, column=1, value=label).font = BOLD_FONT

        base_wacc_scenario = data.get('wacc', 0)
        base_dcf_scenario = data.get('intrinsic_value', 0)

        # Rough sensitivity: DCF ∝ 1/WACC (simplified approximation)
        for col_offset, wacc_delta in enumerate([-0.02, -0.01, 0, 0.01, 0.02], start=0):
            adjusted_wacc = base_wacc_scenario + wacc_delta
            # Simple approximation: intrinsic × (base_wacc / adjusted_wacc)
            adjusted_dcf = base_dcf_scenario * (base_wacc_scenario / adjusted_wacc) if adjusted_wacc > 0 else 0
            cell = ws.cell(row=r, column=2+col_offset, value=adjusted_dcf)
            cell.number_format = NUM_FMT
            if col_offset == 2:  # Base case
                cell.fill = FORMULA_FILL
                cell.font = BOLD_FONT

    # Add note
    r += 2
    ws.cell(row=r, column=1, value='Note: Sensitivity values are approximations using WACC elasticity. For precise sensitivity, re-run full DCF model.')
    ws.cell(row=r, column=1).font = REMARK_FONT
    ws.merge_cells(f'A{r}:H{r}')

    logger.debug(f"Built Beta Scenarios sheet with {len(beta_scenarios)} scenarios")


# ── Sheet 4: DCF Model ─────────────────────────────────────────────────────

def _build_dcf_sheet(ws, result, refs):
    ws.column_dimensions['A'].width = 28
    for c in 'BCDEFGHI':
        ws.column_dimensions[c].width = 16

    dcf = result.get('dcf_details', {})
    projections = dcf.get('fcff_projections', [])
    n_years = len(projections) or 5

    r = 1
    ws.cell(row=r, column=1, value='FCFF-Based DCF Valuation').font = TITLE_FONT

    # ── WACC ──
    r = 3
    _style_section(ws, r, 8, 'WACC Calculation')
    r += 1
    rf_row = r
    ws.cell(row=r, column=1, value='Risk-Free Rate (Rf)')
    _fml(ws, r, 2, f"='Assumptions'!B{refs['rf_rate']}", PCT_FMT)
    _remark(ws, r, 3, 'Ref: Assumptions')

    r += 1
    erp_row = r
    ws.cell(row=r, column=1, value='Equity Risk Premium (ERP)')
    _fml(ws, r, 2, f"='Assumptions'!B{refs['erp']}", PCT_FMT)
    _remark(ws, r, 3, 'Ref: Assumptions')

    r += 1
    beta_row = r
    ws.cell(row=r, column=1, value='Beta (Levered)')
    _fml(ws, r, 2, f"='Assumptions'!B{refs['beta']}", '0.0000')
    _remark(ws, r, 3, 'Ref: Assumptions')

    r += 1
    ke_row = r
    ws.cell(row=r, column=1, value='Cost of Equity (Ke)')
    _fml(ws, r, 2, f'=B{rf_row}+B{beta_row}*B{erp_row}', PCT_FMT)
    _remark(ws, r, 3, '= Rf + Beta * ERP')
    refs['ke'] = ke_row

    r += 1
    kd_row = r
    ws.cell(row=r, column=1, value='Cost of Debt (pre-tax, Kd)')
    _fml(ws, r, 2, f"='Assumptions'!B{refs['cost_of_debt']}", PCT_FMT)

    r += 1
    tax_row = r
    ws.cell(row=r, column=1, value='Tax Rate')
    _fml(ws, r, 2, f"='Assumptions'!B{refs['tax_rate']}", PCT_FMT)

    r += 1
    kd_at_row = r
    ws.cell(row=r, column=1, value='Cost of Debt (after-tax)')
    _fml(ws, r, 2, f'=B{kd_row}*(1-B{tax_row})', PCT_FMT)
    _remark(ws, r, 3, '= Kd * (1 - Tax)')

    r += 1
    dv_row = r
    ws.cell(row=r, column=1, value='Debt Ratio (D/V)')
    _fml(ws, r, 2, f"='Assumptions'!B{refs['debt_ratio']}", PCT_FMT)

    r += 1
    wacc_row = r
    ws.cell(row=r, column=1, value='WACC').font = BOLD_FONT
    _fml(ws, r, 2, f'=B{ke_row}*(1-B{dv_row})+B{kd_at_row}*B{dv_row}', PCT_FMT)
    ws.cell(row=r, column=2).font = BOLD_FONT
    _remark(ws, r, 3, '= Ke*(1-D/V) + Kd_at*(D/V)')
    refs['wacc_row'] = wacc_row

    # ── FCFF Projections ──
    r += 2
    _style_section(ws, r, 2 + n_years, 'FCFF Projections (Rs Cr)')
    r += 1
    header_row = r
    ws.cell(row=r, column=1, value='Metric')
    ws.cell(row=r, column=2, value='Base (FY)')
    for i in range(n_years):
        ws.cell(row=r, column=3 + i, value=f'Year {i+1}')
    _style_header(ws, r, 2 + n_years)

    # Row: Revenue Growth (from Assumptions)
    r += 1
    growth_row = r
    ws.cell(row=r, column=1, value='Revenue Growth')
    for i in range(n_years):
        growth_ref = refs.get(f'growth_y{i+1}')
        if growth_ref:
            _fml(ws, r, 3 + i, f"='Assumptions'!B{growth_ref}", PCT_FMT)
        else:
            _inp(ws, r, 3 + i, projections[i].get('growth_rate', 0.06) if i < len(projections) else 0.06, PCT_FMT)

    # Row: Revenue
    r += 1
    rev_row = r
    ws.cell(row=r, column=1, value='Revenue')
    _fml(ws, r, 2, f"='Assumptions'!B{refs['base_revenue']}", NUM_FMT)
    for i in range(n_years):
        col = 3 + i
        prev_col = get_column_letter(col - 1)
        cur_col = get_column_letter(col)
        _fml(ws, r, col, f'={prev_col}{rev_row}*(1+{cur_col}{growth_row})', NUM_FMT)

    # Row: Margin Damping Factor (full → 0 over projection period)
    r += 1
    damping_row = r
    ws.cell(row=r, column=1, value='Margin Damping Factor')
    damping_factors = [1.0, 0.75, 0.50, 0.25, 0.0]
    for i in range(n_years):
        factor = damping_factors[i] if i < len(damping_factors) else 0.0
        _inp(ws, r, 3 + i, factor, '0.00')
    _remark(ws, r, 2 + n_years + 1, 'Dampens margin improvement: full → 0')

    # Row: EBITDA Margin (dampened: prev + improvement * damping_factor, capped ±3pp)
    r += 1
    margin_row = r
    ws.cell(row=r, column=1, value='EBITDA Margin')
    _fml(ws, r, 2, f"='Assumptions'!B{refs['ebitda_margin']}", PCT_FMT)
    for i in range(n_years):
        col = 3 + i
        prev_col = get_column_letter(col - 1)
        c = get_column_letter(col)
        # Margin = prev_margin + improvement * damping, capped at base ± 3pp
        base_ref = f"'Assumptions'!B{refs['ebitda_margin']}"
        imp_ref = f"'Assumptions'!B{refs['margin_improvement']}"
        _fml(ws, r, col,
             f"=MAX({base_ref}-0.03,MIN({base_ref}+0.03,"
             f"{prev_col}{margin_row}+{imp_ref}*{c}{damping_row}))", PCT_FMT)

    # Row: EBITDA
    r += 1
    ebitda_row = r
    ws.cell(row=r, column=1, value='EBITDA')
    for i in range(n_years):
        col = 3 + i
        c = get_column_letter(col)
        _fml(ws, r, col, f'={c}{rev_row}*{c}{margin_row}', NUM_FMT)

    # Row: Depreciation
    r += 1
    depr_row = r
    ws.cell(row=r, column=1, value='Depreciation')
    for i in range(n_years):
        col = 3 + i
        c = get_column_letter(col)
        _fml(ws, r, col, f"={c}{rev_row}*'Assumptions'!B{refs['depr_to_sales']}", NUM_FMT)

    # Row: EBIT
    r += 1
    ebit_row = r
    ws.cell(row=r, column=1, value='EBIT')
    for i in range(n_years):
        col = 3 + i
        c = get_column_letter(col)
        _fml(ws, r, col, f'={c}{ebitda_row}-{c}{depr_row}', NUM_FMT)

    # Row: Tax Rate
    r += 1
    tax_proj_row = r
    ws.cell(row=r, column=1, value='Tax Rate')
    for i in range(n_years):
        _fml(ws, r, 3 + i, f"='Assumptions'!B{refs['tax_rate']}", PCT_FMT)

    # Row: NOPAT
    r += 1
    nopat_row = r
    ws.cell(row=r, column=1, value='NOPAT')
    for i in range(n_years):
        col = 3 + i
        c = get_column_letter(col)
        _fml(ws, r, col, f'={c}{ebit_row}*(1-{c}{tax_proj_row})', NUM_FMT)

    # Row: Capex
    r += 1
    capex_row = r
    ws.cell(row=r, column=1, value='Capex')
    for i in range(n_years):
        col = 3 + i
        c = get_column_letter(col)
        _fml(ws, r, col, f"={c}{rev_row}*'Assumptions'!B{refs['capex_to_sales']}", NUM_FMT)

    # Row: Delta NWC
    r += 1
    dnwc_row = r
    ws.cell(row=r, column=1, value='Change in NWC')
    for i in range(n_years):
        col = 3 + i
        c = get_column_letter(col)
        prev_c = get_column_letter(col - 1)
        _fml(ws, r, col, f"=({c}{rev_row}-{prev_c}{rev_row})*'Assumptions'!B{refs['nwc_to_sales']}", NUM_FMT)

    # Row: FCFF
    r += 1
    fcff_row = r
    ws.cell(row=r, column=1, value='FCFF').font = BOLD_FONT
    for i in range(n_years):
        col = 3 + i
        c = get_column_letter(col)
        cell = _fml(ws, r, col,
                     f'={c}{nopat_row}+{c}{depr_row}-{c}{capex_row}-{c}{dnwc_row}', NUM_FMT)
        cell.font = BOLD_FONT

    # Row: Discount Factor
    r += 1
    df_row = r
    ws.cell(row=r, column=1, value='Discount Factor')
    for i in range(n_years):
        _fml(ws, r, 3 + i, f'=1/(1+$B${wacc_row})^{i+1}', '0.0000')

    # Row: PV of FCFF
    r += 1
    pv_row = r
    ws.cell(row=r, column=1, value='PV of FCFF')
    for i in range(n_years):
        col = 3 + i
        c = get_column_letter(col)
        _fml(ws, r, col, f'={c}{fcff_row}*{c}{df_row}', NUM_FMT)

    # ── Terminal Value ──
    r += 2
    _style_section(ws, r, 8, 'Terminal Value')
    r += 1
    roce_row = r
    ws.cell(row=r, column=1, value='Terminal ROCE')
    _fml(ws, r, 2, f"='Assumptions'!B{refs['terminal_roce']}", PCT_FMT)

    r += 1
    reinv_row = r
    ws.cell(row=r, column=1, value='Terminal Reinvestment Rate')
    _fml(ws, r, 2, f"='Assumptions'!B{refs['terminal_reinvestment']}", PCT_FMT)

    r += 1
    tg_row = r
    ws.cell(row=r, column=1, value='Terminal Growth (g)')
    _fml(ws, r, 2, f'=MIN(5%,MAX(2%,B{roce_row}*B{reinv_row}))', PCT_FMT)
    _remark(ws, r, 3, '= MIN(5%, MAX(2%, ROCE * Reinvestment))')
    refs['terminal_growth'] = tg_row

    last_fcff_col = get_column_letter(2 + n_years)
    last_rev_col = last_fcff_col
    last_margin_col = last_fcff_col

    # Terminal Revenue
    r += 1
    term_rev_row = r
    ws.cell(row=r, column=1, value='Terminal Revenue')
    _fml(ws, r, 2, f'={last_rev_col}{rev_row}*(1+B{tg_row})', NUM_FMT)
    _remark(ws, r, 3, '= Year5 Revenue * (1 + g)')

    # Terminal EBITDA
    r += 1
    term_ebitda_row = r
    ws.cell(row=r, column=1, value='Terminal EBITDA')
    _fml(ws, r, 2, f'=B{term_rev_row}*{last_margin_col}{margin_row}', NUM_FMT)
    _remark(ws, r, 3, '= Terminal Revenue * Year5 EBITDA Margin')

    # Terminal Depreciation
    r += 1
    term_dep_row = r
    ws.cell(row=r, column=1, value='Terminal Depreciation')
    _fml(ws, r, 2, f"=B{term_rev_row}*'Assumptions'!B{refs['depr_to_sales']}", NUM_FMT)
    _remark(ws, r, 3, '= Terminal Revenue * Dep/Sales')

    # Terminal NOPAT
    r += 1
    term_nopat_row = r
    ws.cell(row=r, column=1, value='Terminal NOPAT')
    _fml(ws, r, 2, f"=(B{term_ebitda_row}-B{term_dep_row})*(1-'Assumptions'!B{refs['tax_rate']})", NUM_FMT)
    _remark(ws, r, 3, '= (EBITDA - Dep) * (1 - Tax)')

    # Terminal FCFF (NOPAT-based)
    r += 1
    fcff_n1_row = r
    ws.cell(row=r, column=1, value='Terminal FCFF')
    _fml(ws, r, 2, f"=B{term_nopat_row}*(1-'Assumptions'!B{refs['terminal_reinvestment']})", NUM_FMT)
    _remark(ws, r, 3, '= NOPAT * (1 - Reinvestment Rate)')

    # Old method comparison row for audit transparency
    r += 1
    old_fcff_row = r
    ws.cell(row=r, column=1, value='Old Method FCFF(n+1)')
    _fml(ws, r, 2, f'={last_fcff_col}{fcff_row}*(1+B{tg_row})', NUM_FMT)
    ws.cell(row=r, column=1).font = REMARK_FONT
    _remark(ws, r, 3, 'Audit: old approach = Year5 FCFF * (1+g)')

    # Terminal Value
    r += 1
    tv_row = r
    ws.cell(row=r, column=1, value='Terminal Value')
    _fml(ws, r, 2, f'=B{fcff_n1_row}/(B{wacc_row}-B{tg_row})', NUM_FMT)
    _remark(ws, r, 3, '= Terminal FCFF / (WACC - g)')

    r += 1
    pv_tv_row = r
    ws.cell(row=r, column=1, value='PV of Terminal Value')
    _fml(ws, r, 2, f'=B{tv_row}/(1+B{wacc_row})^{n_years}', NUM_FMT)

    # ── Equity Bridge ──
    r += 2
    _style_section(ws, r, 8, 'Equity Bridge')
    r += 1
    pv_start = get_column_letter(3)
    pv_end = get_column_letter(2 + n_years)
    pv_explicit_row = r
    ws.cell(row=r, column=1, value='PV of Explicit FCFFs')
    _fml(ws, r, 2, f'=SUM({pv_start}{pv_row}:{pv_end}{pv_row})', NUM_FMT)

    r += 1
    pv_tv_bridge = r
    ws.cell(row=r, column=1, value='PV of Terminal Value')
    _fml(ws, r, 2, f'=B{pv_tv_row}', NUM_FMT)

    r += 1
    fv_row = r
    ws.cell(row=r, column=1, value='Firm (Enterprise) Value').font = BOLD_FONT
    cell = _fml(ws, r, 2, f'=B{pv_explicit_row}+B{pv_tv_bridge}', NUM_FMT)
    cell.font = BOLD_FONT

    r += 1
    debt_bridge = r
    ws.cell(row=r, column=1, value='(-) Net Debt')
    _fml(ws, r, 2, f"='Assumptions'!B{refs['net_debt']}", NUM_FMT)

    r += 1
    cash_bridge = r
    ws.cell(row=r, column=1, value='(+) Cash & Equivalents')
    _fml(ws, r, 2, f"='Assumptions'!B{refs['cash']}", NUM_FMT)

    r += 1
    eq_val_row = r
    ws.cell(row=r, column=1, value='Equity Value').font = BOLD_FONT
    cell = _fml(ws, r, 2, f'=B{fv_row}-B{debt_bridge}+B{cash_bridge}', NUM_FMT)
    cell.font = BOLD_FONT

    r += 1
    shares_bridge = r
    ws.cell(row=r, column=1, value='Shares Outstanding (Cr)')
    _fml(ws, r, 2, f"='Assumptions'!B{refs['shares']}", '0.0000')

    r += 1
    iv_row = r
    ws.cell(row=r, column=1, value='Intrinsic Value per Share (Rs)').font = BOLD_FONT
    cell = _fml(ws, r, 2, f'=B{eq_val_row}/B{shares_bridge}', NUM_FMT)
    cell.font = BOLD_FONT
    _remark(ws, r, 3, '= Equity Value / Shares')
    refs['dcf_iv_row'] = iv_row

    r += 1
    ws.cell(row=r, column=1, value='TV as % of Firm Value')
    _fml(ws, r, 2, f'=B{pv_tv_bridge}/B{fv_row}', PCT_FMT)


# ── Sheet 4: Relative Valuation ─────────────────────────────────────────────

def _build_relative_sheet(ws, result, refs):
    ws.column_dimensions['A'].width = 22
    for c in 'BCDEFGH':
        ws.column_dimensions[c].width = 16

    rel = result.get('relative_details', {})
    implied = rel.get('implied_values', {})
    weights = rel.get('multiple_weights_used', {})
    adj_factors = rel.get('adjustment_factors', {})

    r = 1
    ws.cell(row=r, column=1, value='Relative Valuation').font = TITLE_FONT

    # ── Implied Values ──
    r = 3
    _style_section(ws, r, 6, 'Implied Values by Multiple')
    r += 1
    ws.cell(row=r, column=1, value='Multiple')
    ws.cell(row=r, column=2, value='Peer Median')
    ws.cell(row=r, column=3, value='Company Metric')
    ws.cell(row=r, column=4, value='Implied/Share')
    ws.cell(row=r, column=5, value='Weight')
    ws.cell(row=r, column=6, value='Contribution')
    _style_header(ws, r, 6)

    r += 1
    implied_start = r
    mult_keys = [('pe', 'P/E', 'peer_pe_median'),
                 ('pb', 'P/B', 'peer_pb_median'),
                 ('ev_ebitda', 'EV/EBITDA', 'peer_eveb_median'),
                 ('ps', 'P/S', 'peer_ps_median')]

    for key, label, ref_key in mult_keys:
        imp = implied.get(key, {})
        ws.cell(row=r, column=1, value=label)
        # Peer median refs Assumptions
        _fml(ws, r, 2, f"='Assumptions'!B{refs[ref_key]}", '0.00')
        # Company metric
        _inp(ws, r, 3, imp.get('company_metric', 0), NUM_FMT)
        # Implied/share
        _inp(ws, r, 4, imp.get('implied_per_share', 0), NUM_FMT)
        # Weight
        w = weights.get(key, 0)
        _inp(ws, r, 5, w, PCT_FMT)
        # Contribution = Implied * Weight
        _fml(ws, r, 6, f'=D{r}*E{r}', NUM_FMT)
        r += 1

    implied_end = r - 1

    r += 1
    ws.cell(row=r, column=1, value='Total Weight').font = BOLD_FONT
    total_wt_row = r
    _fml(ws, r, 5, f'=SUM(E{implied_start}:E{implied_end})', PCT_FMT)

    r += 1
    base_val_row = r
    ws.cell(row=r, column=1, value='Base Relative Value').font = BOLD_FONT
    _fml(ws, r, 6, f'=SUM(F{implied_start}:F{implied_end})/E{total_wt_row}', NUM_FMT)

    # ── Quality Adjustments ──
    r += 2
    _style_section(ws, r, 6, 'Quality Adjustments')
    r += 1
    ws.cell(row=r, column=1, value='Factor')
    ws.cell(row=r, column=2, value='Adjustment')
    _style_header(ws, r, 2)

    r += 1
    adj_start = r
    adj_ref_map = {}
    for factor, val in adj_factors.items():
        ws.cell(row=r, column=1, value=factor.replace('_', ' ').title())
        _inp(ws, r, 2, val, PCT_FMT)
        adj_ref_map[factor] = r
        r += 1
    adj_end = r - 1
    refs['adj_factor_rows'] = adj_ref_map

    r += 1
    total_adj_row = r
    ws.cell(row=r, column=1, value='Total Quality Adjustment').font = BOLD_FONT
    if adj_start <= adj_end:
        _fml(ws, r, 2, f'=SUM(B{adj_start}:B{adj_end})', PCT_FMT)
    else:
        _inp(ws, r, 2, 0, PCT_FMT)

    r += 1
    adj_val_row = r
    ws.cell(row=r, column=1, value='Adjusted Relative Value').font = BOLD_FONT
    cell = _fml(ws, r, 2, f'=F{base_val_row}*(1+B{total_adj_row})', NUM_FMT)
    cell.font = BOLD_FONT
    refs['relative_adj_row'] = adj_val_row

    # ── Peer Detail Table ──
    peer_list = rel.get('peer_multiples_list', [])
    if peer_list:
        r += 2
        # Column layout: A=Symbol B=Name C=Tier D=ValGroup E=ValSubgroup
        #   F=CD_Sector G=CD_Industry H=MCap I=P/E J=P/B K=EV/EBITDA L=P/S M=Include
        PEER_COLS = 13
        tight_count = sum(1 for p in peer_list if p.get('tier') == 'tight')
        broad_count = len(peer_list) - tight_count
        _style_section(ws, r, PEER_COLS,
                        f'Peer Detail ({len(peer_list)} peers: {tight_count} tight, {broad_count} broad)')

        # Widen taxonomy columns
        ws.column_dimensions['C'].width = 8
        ws.column_dimensions['D'].width = 16
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 22
        ws.column_dimensions['H'].width = 14
        for col_letter in 'IJKLM':
            ws.column_dimensions[col_letter].width = 12

        r += 1
        peer_header_row = r
        peer_headers = ['Symbol', 'Name', 'Tier', 'Val Group', 'Val Subgroup',
                        'CD Sector', 'CD Industry', 'MCap (Cr)',
                        'P/E', 'P/B', 'EV/EBITDA', 'P/S', 'Include']
        for col_idx, header in enumerate(peer_headers, 1):
            ws.cell(row=r, column=col_idx, value=header)
        _style_header(ws, r, PEER_COLS)

        # Y/N dropdown validation
        yn_dv = DataValidation(type='list', formula1='"Y,N"', allow_blank=False)
        ws.add_data_validation(yn_dv)

        TIGHT_FILL = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
        BROAD_FILL = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')

        r += 1
        peer_start_row = r
        # Sort: tight peers first, then broad
        sorted_peers = sorted(peer_list, key=lambda p: (0 if p.get('tier') == 'tight' else 1,
                                                          -(p.get('mcap') or 0)))
        for peer in sorted_peers:
            tier = peer.get('tier', 'broad')
            row_fill = TIGHT_FILL if tier == 'tight' else BROAD_FILL
            ws.cell(row=r, column=1, value=peer.get('nse_symbol', ''))
            ws.cell(row=r, column=2, value=peer.get('Company Name', ''))
            ws.cell(row=r, column=3, value=tier.title())
            ws.cell(row=r, column=4, value=peer.get('valuation_group', ''))
            ws.cell(row=r, column=5, value=peer.get('valuation_subgroup', ''))
            ws.cell(row=r, column=6, value=peer.get('cd_sector', ''))
            ws.cell(row=r, column=7, value=peer.get('cd_industry', ''))
            ws.cell(row=r, column=8, value=peer.get('mcap', 0)).number_format = INT_FMT
            ws.cell(row=r, column=9, value=peer.get('pe', 0) or 0).number_format = '0.00'
            ws.cell(row=r, column=10, value=peer.get('pb', 0) or 0).number_format = '0.00'
            ws.cell(row=r, column=11, value=peer.get('evebidta', 0) or 0).number_format = '0.00'
            ws.cell(row=r, column=12, value=peer.get('ps', 0) or 0).number_format = '0.00'
            include_cell = ws.cell(row=r, column=13, value='Y')
            include_cell.fill = INPUT_FILL
            yn_dv.add(include_cell)
            # Apply tier shading to info columns
            for c in range(1, 8):
                ws.cell(row=r, column=c).fill = row_fill
            r += 1
        peer_end_row = r - 1
        refs['peer_start'] = peer_start_row
        refs['peer_end'] = peer_end_row

        # ── Filtered Peer Medians (driven by Include Y/N) ──
        # MEDIAN(IF()) array formulas so Assumptions can reference them
        r += 1
        ws.cell(row=r, column=8, value='Filtered Medians →').font = BOLD_FONT
        incl_col = 'M'  # Include column
        pe_col, pb_col, eveb_col, ps_col = 'I', 'J', 'K', 'L'
        for col_letter, col_idx in [(pe_col, 9), (pb_col, 10), (eveb_col, 11), (ps_col, 12)]:
            # Array formula: MEDIAN(IF(Include="Y", metric_range))
            fml = (f'MEDIAN(IF({incl_col}{peer_start_row}:{incl_col}{peer_end_row}="Y",'
                   f'{col_letter}{peer_start_row}:{col_letter}{peer_end_row}))')
            cell = ws.cell(row=r, column=col_idx)
            cell.value = ArrayFormula(ref=f'{col_letter}{r}', text=fml)
            cell.number_format = '0.00'
            cell.font = BOLD_FONT
        filtered_median_row = r
        refs['filtered_median_row'] = filtered_median_row


# ── Sheet 5: Monte Carlo ────────────────────────────────────────────────────

def _build_monte_carlo_sheet(ws, result, refs):
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 18

    r = 1
    ws.cell(row=r, column=1, value='Monte Carlo Simulation').font = TITLE_FONT

    mc_sims = int(os.getenv('MC_SIMULATIONS', 10000))
    r += 2
    ws.cell(row=r, column=1, value='Number of Simulations')
    ws.cell(row=r, column=2, value=mc_sims).number_format = INT_FMT

    r += 2
    _style_section(ws, r, 2, 'Distribution Statistics')
    r += 1
    for label, key in [('Mean', 'mc_mean'), ('Median', 'mc_median')]:
        ws.cell(row=r, column=1, value=label)
        val = result.get(key, 0) or 0
        ws.cell(row=r, column=2, value=val).number_format = NUM_FMT
        if key == 'mc_median':
            refs['mc_median_row'] = r
        r += 1

    percentiles = result.get('mc_percentiles', {})
    for label in ['5th', '10th', '25th', '75th', '90th', '95th']:
        ws.cell(row=r, column=1, value=f'{label} Percentile')
        ws.cell(row=r, column=2, value=percentiles.get(label, 0) or 0).number_format = NUM_FMT
        r += 1

    r += 1
    _style_section(ws, r, 2, 'Market Comparison')
    r += 1
    cmp = result.get('cmp', 0)
    ws.cell(row=r, column=1, value='Current Market Price')
    ws.cell(row=r, column=2, value=cmp).number_format = NUM_FMT
    r += 1
    ws.cell(row=r, column=1, value='P(Value > CMP)')
    ws.cell(row=r, column=2, value=result.get('mc_probability_above_cmp', 0) or 0).number_format = PCT_FMT

    r += 2
    _style_section(ws, r, 2, 'Randomized Parameters')
    r += 1
    ws.cell(row=r, column=1, value='Parameter')
    ws.cell(row=r, column=2, value='Distribution')
    _style_header(ws, r, 2)
    r += 1
    for name, dist in [('Revenue Growth', 'Triangular: base * [0.7, 1.0, 1.3]'),
                        ('EBITDA Margin', 'Normal: mean=base, std=10% of base'),
                        ('Terminal ROCE', 'Normal: mean=base, std=15% of base'),
                        ('Beta', 'Normal: mean=base, std=0.15')]:
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=dist)
        r += 1


# ── Sheet 6: Sensitivity ────────────────────────────────────────────────────

def _build_sensitivity_sheet(ws, result, refs):
    ws.column_dimensions['A'].width = 18

    sensitivity = result.get('sensitivity', {})
    table = sensitivity.get('sensitivity_table', [])
    wacc_values = sensitivity.get('wacc_values', [])
    growth_values = sensitivity.get('growth_values', [])

    r = 1
    ws.cell(row=r, column=1, value='Sensitivity Analysis').font = TITLE_FONT
    r += 1
    ws.cell(row=r, column=1, value='Intrinsic value (Rs) for WACC vs Terminal Growth')

    if not table:
        r += 2
        ws.cell(row=r, column=1, value='No sensitivity data available')
        return

    base_value = sensitivity.get('base_intrinsic', 0)
    r += 2
    ws.cell(row=r, column=1, value=f'Base intrinsic: Rs {base_value:,.2f}')

    r += 2
    ws.cell(row=r, column=1, value='WACC \\ Growth').font = BOLD_FONT
    for j, g in enumerate(growth_values):
        cell = ws.cell(row=r, column=2 + j, value=g)
        cell.number_format = PCT_FMT
        cell.font = BOLD_FONT
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[get_column_letter(2 + j)].width = 14
    _style_header(ws, r, 1 + len(growth_values))

    for i, wacc_val in enumerate(wacc_values):
        r += 1
        ws.cell(row=r, column=1, value=wacc_val).number_format = PCT_FMT
        ws.cell(row=r, column=1).font = BOLD_FONT
        if i < len(table):
            for j, val in enumerate(table[i]):
                cell = ws.cell(row=r, column=2 + j, value=val)
                cell.number_format = NUM_FMT
                if abs(val - base_value) < 1:
                    cell.fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')
                elif val > base_value * 1.2:
                    cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                elif val < base_value * 0.8:
                    cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')


# ── Sheet 7: Macro Drivers (NEW) ────────────────────────────────────────────

def _build_macro_drivers_sheet(ws, result, refs):
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 30

    hierarchy = result.get('driver_hierarchy', {})

    r = 1
    ws.cell(row=r, column=1, value=f'Macro Drivers (Weight: {hierarchy.get("macro_weight", 0.15):.0%})').font = TITLE_FONT
    r += 1
    ws.cell(row=r, column=1, value=hierarchy.get('principle', 'Macro sets the ceiling and floor.')).font = REMARK_FONT

    r += 2
    ws.cell(row=r, column=1, value='Parameter')
    ws.cell(row=r, column=2, value='Value')
    ws.cell(row=r, column=3, value='Source')
    ws.cell(row=r, column=4, value='Feeds Into')
    _style_header(ws, r, 4)

    r += 1
    ws.cell(row=r, column=1, value='Risk-Free Rate (Rf)')
    _fml(ws, r, 2, f"='Assumptions'!B{refs['rf_rate']}", PCT_FMT)
    ws.cell(row=r, column=3, value='Macro: interest_rate_10y')
    ws.cell(row=r, column=4, value='WACC via Cost of Equity (Ke)')

    r += 1
    ws.cell(row=r, column=1, value='Equity Risk Premium (ERP)')
    _fml(ws, r, 2, f"='Assumptions'!B{refs['erp']}", PCT_FMT)
    ws.cell(row=r, column=3, value='India ERP')
    ws.cell(row=r, column=4, value='WACC via Cost of Equity (Ke)')

    r += 1
    ws.cell(row=r, column=1, value='India 10Y Yield')
    assumptions = result.get('dcf_assumptions', {})
    ws.cell(row=r, column=2, value=assumptions.get('risk_free_rate', 0.0674)).number_format = PCT_FMT
    ws.cell(row=r, column=3, value='Macro: market_indicators.csv')
    ws.cell(row=r, column=4, value='Rf calibration')

    r += 2
    _style_section(ws, r, 4, 'How Macro Flows to Valuation')
    r += 1
    ws.cell(row=r, column=1, value='Ke = Rf + Beta * ERP')
    if refs.get('ke'):
        _fml(ws, r, 2, f"='DCF Model'!B{refs['ke']}", PCT_FMT)
    r += 1
    ws.cell(row=r, column=1, value='WACC = Ke*(1-D/V) + Kd_at*(D/V)')
    if refs.get('wacc_row'):
        _fml(ws, r, 2, f"='DCF Model'!B{refs['wacc_row']}", PCT_FMT)
    r += 1
    ws.cell(row=r, column=1, value='Terminal Growth bounded by GDP')
    if refs.get('terminal_growth'):
        _fml(ws, r, 2, f"='DCF Model'!B{refs['terminal_growth']}", PCT_FMT)


# ── Sheet 8: Sector Drivers (NEW) ───────────────────────────────────────────

def _build_sector_drivers_sheet(ws, result, refs):
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 16

    group_cfg = result.get('sector_drivers_config', {})
    hierarchy = result.get('driver_hierarchy', {})
    outlook = result.get('sector_outlook', {})
    sector_name = group_cfg.get('csv_sector_name', result.get('sector', ''))
    positives = set(outlook.get('key_positives', []))
    negatives = set(outlook.get('key_negatives', []))

    r = 1
    ws.cell(row=r, column=1,
            value=f'Group Drivers: {sector_name} (Weight: {hierarchy.get("group_weight", 0.20):.0%})').font = TITLE_FONT

    r += 2
    ws.cell(row=r, column=1, value='Category')
    ws.cell(row=r, column=2, value='Driver')
    ws.cell(row=r, column=3, value='Weight')
    ws.cell(row=r, column=4, value='Direction')
    ws.cell(row=r, column=5, value='Weighted Score')
    _style_header(ws, r, 5)

    # Direction dropdown
    dir_dv = DataValidation(type='list', formula1='"POSITIVE,NEUTRAL,NEGATIVE"',
                            allow_blank=False)
    ws.add_data_validation(dir_dv)

    r += 1
    driver_start = r
    has_drivers = False
    for category in ['demand_drivers', 'cost_drivers', 'regulatory_drivers',
                     'group_specific_drivers']:
        drivers = group_cfg.get(category, [])
        cat_label = category.replace('_', ' ').title().replace('Drivers', '').strip()
        for driver in drivers:
            has_drivers = True
            name = driver.get('name', '')
            weight = driver.get('weight', 0)
            # Set initial direction from outlook
            if name in positives:
                direction = 'POSITIVE'
            elif name in negatives:
                direction = 'NEGATIVE'
            else:
                direction = 'NEUTRAL'

            ws.cell(row=r, column=1, value=cat_label)
            ws.cell(row=r, column=2, value=name)
            ws.cell(row=r, column=3, value=weight).number_format = '0.00'
            dir_cell = _inp(ws, r, 4, direction)
            dir_dv.add(dir_cell)
            # Weighted score formula
            _fml(ws, r, 5,
                 f'=IF(D{r}="POSITIVE",1,IF(D{r}="NEGATIVE",-1,0))*C{r}', '0.0000')
            r += 1

    # Handle empty driver list
    if not has_drivers:
        ws.cell(row=r, column=1, value='No drivers configured')
        ws.cell(row=r, column=1).font = REMARK_FONT
        r += 1

    driver_end = max(r - 1, driver_start)

    # Totals
    r += 1
    ws.cell(row=r, column=2, value='Total Weight').font = BOLD_FONT
    total_wt_row = r
    if has_drivers:
        _fml(ws, r, 3, f'=SUM(C{driver_start}:C{driver_end})', '0.00')
    else:
        ws.cell(row=r, column=3, value=0).number_format = '0.00'

    r += 1
    ws.cell(row=r, column=2, value='Outlook Score').font = BOLD_FONT
    score_row = r
    if has_drivers:
        # Use IFERROR to avoid division by zero
        _fml(ws, r, 5, f'=IFERROR(SUM(E{driver_start}:E{driver_end})/C{total_wt_row},0)', '0.0000')
    else:
        ws.cell(row=r, column=5, value=0).number_format = '0.0000'

    r += 2
    _style_section(ws, r, 5, 'Outlook to Valuation Adjustment')
    r += 1
    ws.cell(row=r, column=1, value='Growth Adjustment')
    _fml(ws, r, 2, f'=E{score_row}*0.03', PCT_FMT)
    _remark(ws, r, 3, 'max +/-3%')
    r += 1
    ws.cell(row=r, column=1, value='Margin Adjustment')
    _fml(ws, r, 2, f'=E{score_row}*0.01', PCT_FMT)
    _remark(ws, r, 3, 'max +/-1%')

    # Porter's Five Forces
    porter = group_cfg.get('porter_forces', {})
    if porter:
        r += 2
        _style_section(ws, r, 5, "Porter's Five Forces")
        r += 1
        ws.cell(row=r, column=1, value='Force')
        ws.cell(row=r, column=2, value='Level')
        ws.cell(row=r, column=3, value='Impact')
        _style_header(ws, r, 3)
        r += 1
        for force_key in ['entry_barriers', 'supplier_power', 'buyer_power',
                          'substitutes', 'rivalry']:
            force = porter.get(force_key, {})
            if isinstance(force, dict):
                ws.cell(row=r, column=1, value=force_key.replace('_', ' ').title())
                ws.cell(row=r, column=2, value=force.get('value', ''))
                ws.cell(row=r, column=3, value=force.get('impact', ''))
                r += 1


# ── Sheet 8b: Subgroup Drivers (NEW for 4-level hierarchy) ──────────────────

def _build_subgroup_drivers_sheet(ws, result, refs):
    """Build the Subgroup Drivers sheet (35% weight in 4-level hierarchy)."""
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 16

    hierarchy = result.get('driver_hierarchy', {})
    valuation_subgroup = result.get('valuation_subgroup', '')
    valuation_group = result.get('valuation_group', '')

    r = 1
    ws.cell(row=r, column=1,
            value=f'Subgroup Drivers: {valuation_subgroup or "(none)"} (Weight: {hierarchy.get("subgroup_weight", 0.35):.0%})'
    ).font = TITLE_FONT

    r += 1
    ws.cell(row=r, column=1, value=f'Parent Group: {valuation_group}').font = REMARK_FONT

    r += 2
    # Headers
    headers = ['Driver', 'Category', 'Current State', 'Trend', 'Impact']
    for c, h in enumerate(headers, 1):
        ws.cell(row=r, column=c, value=h).font = HEADER_FONT
        ws.cell(row=r, column=c).fill = HEADER_FILL
    r += 1

    # Subgroup-specific drivers
    subgroup_drivers = result.get('subgroup_drivers', {})
    sector_outlook = result.get('sector_outlook', {})

    # If subgroup_drivers is empty, try to get from sector_outlook
    if not subgroup_drivers and sector_outlook:
        subgroup_drivers = sector_outlook.get('subgroup_drivers', {})

    driver_start = r
    if subgroup_drivers:
        for driver_key, driver in subgroup_drivers.items():
            if isinstance(driver, dict):
                ws.cell(row=r, column=1, value=driver.get('name', driver_key.replace('SUBGROUP_', '')))
                ws.cell(row=r, column=2, value=(driver.get('category') or '').replace('_', ' ').title())
                ws.cell(row=r, column=3, value=str(driver.get('value', 'N/A')))
                ws.cell(row=r, column=4, value=driver.get('trend', 'STABLE'))
                ws.cell(row=r, column=5, value=driver.get('direction', driver.get('impact_direction', 'NEUTRAL')))
                r += 1
    else:
        ws.cell(row=r, column=1, value='No subgroup drivers defined')
        ws.cell(row=r, column=1).font = REMARK_FONT
        r += 1
    driver_end = r - 1

    # Summary
    r += 2
    _style_section(ws, r, 5, 'Subgroup Outlook')
    r += 1
    ws.cell(row=r, column=1, value='Subgroup Score')
    subgroup_score = result.get('subgroup_outlook', sector_outlook.get('subgroup_score', 0))
    ws.cell(row=r, column=2, value=subgroup_score).number_format = '0.0000'
    r += 1
    ws.cell(row=r, column=1, value='Combined Score')
    ws.cell(row=r, column=2, value=sector_outlook.get('outlook_score', 0)).number_format = '0.0000'

    r += 2
    _style_section(ws, r, 5, 'Role in Valuation')
    r += 1
    ws.cell(row=r, column=1, value='Subgroups differentiate companies within the same Group.')
    ws.cell(row=r, column=1).font = REMARK_FONT
    r += 1
    ws.cell(row=r, column=1, value='E.g., INDUSTRIALS_DEFENSE vs INDUSTRIALS_ENGINEERING')
    ws.cell(row=r, column=1).font = REMARK_FONT


# ── Sheet 9: Company Drivers (NEW) ──────────────────────────────────────────

def _build_company_drivers_sheet(ws, result, refs):
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 35

    company_cfg = result.get('company_alpha_config', {})
    hierarchy = result.get('driver_hierarchy', {})
    company_name = company_cfg.get('csv_name', result.get('company_name', ''))

    r = 1
    ws.cell(row=r, column=1,
            value=f'Company Drivers: {company_name} (Weight: {hierarchy.get("company_weight", 0.30):.0%})').font = TITLE_FONT

    # Alpha thesis
    thesis = company_cfg.get('alpha_thesis', {})
    if thesis:
        r += 1
        ws.cell(row=r, column=1, value='Bull Thesis').font = BOLD_FONT
        ws.cell(row=r, column=2, value=thesis.get('bull', ''))
        r += 1
        ws.cell(row=r, column=1, value='Bear Thesis').font = BOLD_FONT
        ws.cell(row=r, column=2, value=thesis.get('bear', ''))
        r += 1
        ws.cell(row=r, column=1, value='Key Moat').font = BOLD_FONT
        ws.cell(row=r, column=2, value=thesis.get('key_moat', ''))

    # Alpha drivers
    alpha_drivers = company_cfg.get('alpha_drivers', {})
    if alpha_drivers:
        r += 2
        ws.cell(row=r, column=1, value='Category')
        ws.cell(row=r, column=2, value='Driver')
        ws.cell(row=r, column=3, value='Weight')
        ws.cell(row=r, column=4, value='How It Maps to Valuation')
        _style_header(ws, r, 4)

        r += 1
        driver_start = r
        impact_map = {
            'market_share': 'ROCE premium in Relative Val',
            'mgmt_quality': 'ROCE, terminal reinvestment, margin',
            'capex': 'Revenue growth, Capex/Sales ratio',
            'geographic_mix': 'Revenue diversification, FX',
            'product_mix': 'Margin, ASP growth',
            'balance_sheet': 'Debt ratio, cash position',
            'governance': 'Governance discount in Relative Val',
            'key_personnel': 'Execution risk, confidence score',
            'litigation': 'Risk discount',
        }

        for category, drivers in alpha_drivers.items():
            cat_label = category.replace('_', ' ').title()
            if isinstance(drivers, list):
                for driver in drivers:
                    ws.cell(row=r, column=1, value=cat_label)
                    ws.cell(row=r, column=2, value=driver.get('name', ''))
                    ws.cell(row=r, column=3, value=driver.get('weight', 0)).number_format = '0.00'
                    ws.cell(row=r, column=4, value=impact_map.get(category, ''))
                    r += 1
        driver_end = r - 1

        r += 1
        ws.cell(row=r, column=2, value='Total Alpha Weight').font = BOLD_FONT
        _fml(ws, r, 3, f'=SUM(C{driver_start}:C{driver_end})', '0.00')

    # Quality Adjustments cross-reference
    adj_rows = refs.get('adj_factor_rows', {})
    if adj_rows:
        r += 2
        _style_section(ws, r, 4, 'Quality Adjustments (from Relative Val)')
        r += 1
        for factor, adj_row in adj_rows.items():
            ws.cell(row=r, column=1, value=factor.replace('_', ' ').title())
            _fml(ws, r, 2, f"='Relative Val'!B{adj_row}", PCT_FMT)
            r += 1
        ws.cell(row=r, column=1, value='Net Quality Adjustment').font = BOLD_FONT
        if refs.get('relative_adj_row'):
            _remark(ws, r, 2, f"See 'Relative Val' adjusted value")


# ── Consolidated Driver Hierarchy Sheet ──────────────────────────────────────

def _build_driver_hierarchy_sheet(ws, result, refs):
    """
    Consolidated 4-level driver hierarchy in one sheet:
    - MACRO (15%): Interest rates, GDP, currency
    - GROUP (20%): Industry cycle, structural trends
    - SUBGROUP (35%): Key differentiators
    - COMPANY (30%): Alpha thesis, execution
    """
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 30

    hierarchy = result.get('driver_hierarchy', {})
    outlook = result.get('sector_outlook', {})
    group_cfg = result.get('sector_drivers_config', {})  # Fallback to config
    group_drivers = result.get('group_drivers', outlook.get('group_drivers', {}))
    company_cfg = result.get('company_alpha_config', {})

    r = 1
    ws.cell(row=r, column=1, value='Driver Hierarchy (4-Level)').font = TITLE_FONT
    r += 1
    ws.cell(row=r, column=1,
            value=hierarchy.get('principle', 'Macro sets ceiling. Group sets cycle. Subgroup differentiates. Company delivers alpha.')
    ).font = REMARK_FONT

    # ── MACRO LEVEL (15%) ──
    r += 2
    _style_section(ws, r, 6, f'MACRO LEVEL (Weight: {hierarchy.get("macro_weight", 0.15):.0%})')
    r += 1
    ws.cell(row=r, column=1, value='Parameter')
    ws.cell(row=r, column=2, value='Value')
    ws.cell(row=r, column=3, value='Source')
    _style_header(ws, r, 3)

    r += 1
    assumptions = result.get('dcf_assumptions', {})
    ws.cell(row=r, column=1, value='Risk-Free Rate (Rf)')
    ws.cell(row=r, column=2, value=assumptions.get('risk_free_rate', 0.0674)).number_format = PCT_FMT
    ws.cell(row=r, column=3, value='Macro: interest_rate_10y')
    r += 1
    ws.cell(row=r, column=1, value='Equity Risk Premium (ERP)')
    ws.cell(row=r, column=2, value=assumptions.get('erp', 0.0708)).number_format = PCT_FMT
    ws.cell(row=r, column=3, value='India ERP')
    r += 1
    ws.cell(row=r, column=1, value='Beta (Levered)')
    ws.cell(row=r, column=2, value=assumptions.get('beta', 1.0)).number_format = '0.00'
    ws.cell(row=r, column=3, value='Sector beta')

    # ── GROUP LEVEL (20%) ──
    r += 2
    valuation_group = result.get('valuation_group', group_cfg.get('csv_sector_name', ''))
    _style_section(ws, r, 6, f'GROUP LEVEL: {valuation_group} (Weight: {hierarchy.get("group_weight", 0.20):.0%})')
    r += 1
    ws.cell(row=r, column=1, value='Category')
    ws.cell(row=r, column=2, value='Driver')
    ws.cell(row=r, column=3, value='Weight')
    ws.cell(row=r, column=4, value='Direction')
    ws.cell(row=r, column=5, value='Score')
    _style_header(ws, r, 5)

    r += 1
    positives = set(outlook.get('key_positives', []))
    negatives = set(outlook.get('key_negatives', []))
    group_driver_start = r
    has_group_drivers = False

    # Use loaded group_drivers from DB/outlook, fallback to config
    if group_drivers:
        for driver_key, driver in group_drivers.items():
            has_group_drivers = True
            name = driver.get('name', driver_key.replace('GROUP_', ''))
            weight = driver.get('weight', 0)
            category = driver.get('category') or 'DEMAND'
            direction = driver.get('impact_direction', 'NEUTRAL')

            ws.cell(row=r, column=1, value=category.title())
            ws.cell(row=r, column=2, value=name)
            ws.cell(row=r, column=3, value=weight).number_format = '0.00'
            ws.cell(row=r, column=4, value=direction)
            score = weight if direction == 'POSITIVE' else (-weight if direction == 'NEGATIVE' else 0)
            ws.cell(row=r, column=5, value=score).number_format = '0.00'
            r += 1
    else:
        # Fallback to config-based drivers
        for category in ['demand_drivers', 'cost_drivers', 'regulatory_drivers', 'group_specific_drivers']:
            drivers = group_cfg.get(category, [])
            cat_label = category.replace('_', ' ').title().replace('Drivers', '').strip()
            for driver in drivers:
                has_group_drivers = True
                name = driver.get('name', '')
                weight = driver.get('weight', 0)
                direction = 'POSITIVE' if name in positives else ('NEGATIVE' if name in negatives else 'NEUTRAL')

                ws.cell(row=r, column=1, value=cat_label)
                ws.cell(row=r, column=2, value=name)
                ws.cell(row=r, column=3, value=weight).number_format = '0.00'
                ws.cell(row=r, column=4, value=direction)
                score = weight if direction == 'POSITIVE' else (-weight if direction == 'NEGATIVE' else 0)
                ws.cell(row=r, column=5, value=score).number_format = '0.00'
                r += 1

    if not has_group_drivers:
        ws.cell(row=r, column=1, value='No group drivers configured').font = REMARK_FONT
        r += 1

    r += 1
    ws.cell(row=r, column=1, value='Group Outlook Score').font = BOLD_FONT
    ws.cell(row=r, column=5, value=outlook.get('group_score', 0)).number_format = '0.0000'

    # ── SUBGROUP LEVEL (35%) ──
    r += 2
    valuation_subgroup = result.get('valuation_subgroup', '')
    _style_section(ws, r, 6, f'SUBGROUP LEVEL: {valuation_subgroup or "(none)"} (Weight: {hierarchy.get("subgroup_weight", 0.35):.0%})')
    r += 1
    ws.cell(row=r, column=1, value='Driver')
    ws.cell(row=r, column=2, value='Category')
    ws.cell(row=r, column=3, value='State')
    ws.cell(row=r, column=4, value='Trend')
    ws.cell(row=r, column=5, value='Impact')
    _style_header(ws, r, 5)

    r += 1
    subgroup_drivers = result.get('subgroup_drivers', outlook.get('subgroup_drivers', {}))
    if subgroup_drivers:
        for key, driver in subgroup_drivers.items():
            if isinstance(driver, dict):
                ws.cell(row=r, column=1, value=driver.get('name', key.replace('SUBGROUP_', '')))
                ws.cell(row=r, column=2, value=(driver.get('category') or '').replace('_', ' ').title())
                ws.cell(row=r, column=3, value=str(driver.get('value', 'N/A')))
                ws.cell(row=r, column=4, value=driver.get('trend', 'STABLE'))
                ws.cell(row=r, column=5, value=driver.get('direction', driver.get('impact_direction', 'NEUTRAL')))
                r += 1
    else:
        ws.cell(row=r, column=1, value='No subgroup drivers defined').font = REMARK_FONT
        r += 1

    r += 1
    ws.cell(row=r, column=1, value='Subgroup Score').font = BOLD_FONT
    ws.cell(row=r, column=5, value=result.get('subgroup_outlook', outlook.get('subgroup_score', 0))).number_format = '0.0000'

    # ── COMPANY LEVEL (30%) ──
    r += 2
    company_name = company_cfg.get('csv_name', result.get('company_name', ''))
    _style_section(ws, r, 6, f'COMPANY LEVEL: {company_name} (Weight: {hierarchy.get("company_weight", 0.30):.0%})')

    company_drivers = result.get('company_drivers', {})
    company_adj = result.get('company_adjustment', {})

    if company_drivers:
        # Split into auto and qualitative
        auto_drivers = {k: v for k, v in company_drivers.items()
                        if isinstance(v, dict) and v.get('source') == 'AUTO'}
        qual_drivers = {k: v for k, v in company_drivers.items()
                        if isinstance(v, dict) and v.get('source') != 'AUTO'}

        # Auto-computed drivers table
        if auto_drivers:
            r += 1
            ws.cell(row=r, column=1, value='Auto-Computed Drivers (affect growth + margin)').font = BOLD_FONT
            r += 1
            ws.cell(row=r, column=1, value='Driver Name')
            ws.cell(row=r, column=2, value='Category')
            ws.cell(row=r, column=3, value='Current Value')
            ws.cell(row=r, column=4, value='Direction')
            ws.cell(row=r, column=5, value='Trend')
            ws.cell(row=r, column=6, value='Weight')
            _style_header(ws, r, 6)

            r += 1
            for key, driver in sorted(auto_drivers.items()):
                ws.cell(row=r, column=1, value=driver.get('name', key.replace('COMPANY_', '')))
                ws.cell(row=r, column=2, value=(driver.get('category') or '').replace('_', ' ').title())
                ws.cell(row=r, column=3, value=str(driver.get('value', 'N/A')))
                ws.cell(row=r, column=4, value=driver.get('impact_direction', 'NEUTRAL'))
                ws.cell(row=r, column=5, value=driver.get('trend', 'STABLE'))
                ws.cell(row=r, column=6, value=driver.get('weight', 0)).number_format = '0.0000'
                r += 1

        # PM-curated qualitative drivers table
        if qual_drivers:
            r += 1
            ws.cell(row=r, column=1, value='PM-Curated Qualitative Drivers (affect terminal ROCE/reinvestment)').font = BOLD_FONT
            r += 1
            ws.cell(row=r, column=1, value='Driver Name')
            ws.cell(row=r, column=2, value='Category')
            ws.cell(row=r, column=3, value='Current Value')
            ws.cell(row=r, column=4, value='Direction')
            ws.cell(row=r, column=5, value='Trend')
            ws.cell(row=r, column=6, value='Weight')
            _style_header(ws, r, 6)

            r += 1
            for key, driver in sorted(qual_drivers.items()):
                ws.cell(row=r, column=1, value=driver.get('name', key.replace('COMPANY_', '')))
                ws.cell(row=r, column=2, value=(driver.get('category') or '').replace('_', ' ').title())
                val = driver.get('value')
                ws.cell(row=r, column=3, value=str(val) if val else 'Not set')
                direction = driver.get('impact_direction', 'NEUTRAL')
                ws.cell(row=r, column=4, value=direction)
                ws.cell(row=r, column=5, value=driver.get('trend', 'STABLE'))
                ws.cell(row=r, column=6, value=driver.get('weight', 0)).number_format = '0.0000'
                # Highlight if NEUTRAL (PM hasn't set a view)
                if direction == 'NEUTRAL':
                    ws.cell(row=r, column=4).font = REMARK_FONT
                r += 1

        # Adjustments Applied
        r += 1
        ws.cell(row=r, column=1, value='Adjustments Applied').font = BOLD_FONT
        r += 1
        ws.cell(row=r, column=1, value='Growth Adjustment (auto drivers)')
        ws.cell(row=r, column=3, value=company_adj.get('growth_adj', 0)).number_format = '0.00%'
        r += 1
        ws.cell(row=r, column=1, value='Margin Adjustment (auto drivers)')
        ws.cell(row=r, column=3, value=company_adj.get('margin_adj', 0)).number_format = '0.00%'
        r += 1
        ws.cell(row=r, column=1, value='Terminal ROCE Adjustment (PM qualitative)')
        ws.cell(row=r, column=3, value=company_adj.get('terminal_roce_adj', 0)).number_format = '0.00%'
        r += 1
        ws.cell(row=r, column=1, value='Terminal Reinvestment Adjustment (PM qualitative)')
        ws.cell(row=r, column=3, value=company_adj.get('terminal_reinv_adj', 0)).number_format = '0.00%'

    else:
        r += 1
        ws.cell(row=r, column=1, value='No company drivers populated yet').font = REMARK_FONT

    r += 1
    ws.cell(row=r, column=1, value='Company Score').font = BOLD_FONT
    ws.cell(row=r, column=5, value=company_adj.get('company_score', 0)).number_format = '0.0000'

    # ── COMBINED OUTLOOK ──
    r += 2
    _style_section(ws, r, 6, 'COMBINED OUTLOOK')
    r += 1
    ws.cell(row=r, column=1, value='Level')
    ws.cell(row=r, column=2, value='Weight')
    ws.cell(row=r, column=3, value='Score')
    ws.cell(row=r, column=4, value='Contribution')
    _style_header(ws, r, 4)

    r += 1
    ws.cell(row=r, column=1, value='Macro')
    ws.cell(row=r, column=2, value=hierarchy.get('macro_weight', 0.15)).number_format = PCT_FMT
    ws.cell(row=r, column=3, value='(embedded in WACC)')
    r += 1
    ws.cell(row=r, column=1, value='Group')
    ws.cell(row=r, column=2, value=hierarchy.get('group_weight', 0.20)).number_format = PCT_FMT
    ws.cell(row=r, column=3, value=outlook.get('group_score', 0)).number_format = '0.0000'
    r += 1
    ws.cell(row=r, column=1, value='Subgroup')
    ws.cell(row=r, column=2, value=hierarchy.get('subgroup_weight', 0.35)).number_format = PCT_FMT
    ws.cell(row=r, column=3, value=result.get('subgroup_outlook', outlook.get('subgroup_score', 0))).number_format = '0.0000'
    r += 1
    ws.cell(row=r, column=1, value='Company')
    ws.cell(row=r, column=2, value=hierarchy.get('company_weight', 0.30)).number_format = PCT_FMT
    company_score = result.get('company_adjustment', {}).get('company_score', 0)
    ws.cell(row=r, column=3, value=company_score).number_format = '0.0000'

    r += 1
    ws.cell(row=r, column=1, value='Combined Outlook').font = BOLD_FONT
    ws.cell(row=r, column=3, value=outlook.get('outlook_score', 0)).number_format = '0.0000'
    ws.cell(row=r, column=4, value=outlook.get('outlook_label', 'NEUTRAL')).font = BOLD_FONT


# ── Computation Log Sheet ───────────────────────────────────────────────────

def _build_log_sheet(ws, result, refs):
    """Build the Computation Log sheet with timestamped log entries."""
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 120

    r = 1
    ws.cell(row=r, column=1, value='Computation Log').font = TITLE_FONT
    r += 1
    ws.cell(row=r, column=1,
            value=f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}').font = REMARK_FONT

    r += 2
    ws.cell(row=r, column=1, value='Timestamp')
    ws.cell(row=r, column=2, value='Module')
    ws.cell(row=r, column=3, value='Level')
    ws.cell(row=r, column=4, value='Message')
    _style_header(ws, r, 4)

    log_records = result.get('_computation_logs', [])
    r += 1
    for record in log_records:
        # Timestamp
        if hasattr(record, 'created'):
            ts = datetime.fromtimestamp(record.created).strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]
        else:
            ts = ''
        ws.cell(row=r, column=1, value=ts).font = LOG_FONT

        # Module
        module = getattr(record, 'name', '') if hasattr(record, 'name') else ''
        ws.cell(row=r, column=2, value=module).font = LOG_FONT

        # Level
        level = getattr(record, 'levelname', '') if hasattr(record, 'levelname') else ''
        level_cell = ws.cell(row=r, column=3, value=level)
        level_cell.font = LOG_FONT

        # Message
        msg = getattr(record, 'getMessage', lambda: '')() if hasattr(record, 'getMessage') else str(record)
        msg_cell = ws.cell(row=r, column=4, value=msg[:2000])  # Truncate long messages
        msg_cell.font = LOG_FONT

        # Color-code by level
        if level == 'WARNING':
            for col in range(1, 5):
                ws.cell(row=r, column=col).fill = WARNING_FILL
        elif level == 'ERROR' or level == 'CRITICAL':
            for col in range(1, 5):
                ws.cell(row=r, column=col).fill = ERROR_FILL
        elif level == 'DEBUG':
            for col in range(1, 5):
                ws.cell(row=r, column=col).font = DEBUG_FONT

        r += 1

    # Freeze header row
    ws.freeze_panes = 'A5'

    if not log_records:
        ws.cell(row=r, column=1, value='No computation logs captured').font = REMARK_FONT


# ── Sheet 1: Summary ────────────────────────────────────────────────────────

def _build_summary_sheet(ws, result, refs):
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18

    company = result.get('company_name', '')
    symbol = result.get('nse_symbol', '')

    r = 1
    ws.cell(row=r, column=1, value=f'{company} ({symbol}) - Valuation Report').font = TITLE_FONT
    r += 1
    ws.cell(row=r, column=1, value=f'Date: {result.get("valuation_date", date.today().isoformat())}')
    valuation_group = result.get('valuation_group', result.get('sector', ''))
    valuation_subgroup = result.get('valuation_subgroup', '')
    ws.cell(row=r, column=2, value=f'Group: {valuation_group}')
    ws.cell(row=r, column=3, value=f'Subgroup: {valuation_subgroup}' if valuation_subgroup else '')

    # Market data
    r += 2
    _style_section(ws, r, 5, 'Market Data')
    r += 1
    cmp_row = r
    # CMP with date from daily_date column in core CSV
    price_date = result.get('price_date', result.get('daily_date', ''))
    ws.cell(row=r, column=1, value='Current Market Price (CMP)')
    ws.cell(row=r, column=2, value=result.get('cmp', 0)).number_format = NUM_FMT
    ws.cell(row=r, column=3, value=f'as of {price_date}' if price_date else '').font = REMARK_FONT
    r += 1
    ws.cell(row=r, column=1, value='Market Cap (Rs Cr)')
    ws.cell(row=r, column=2, value=result.get('mcap_cr')).number_format = NUM_FMT

    # Valuation verdict
    r += 2
    _style_section(ws, r, 5, 'Valuation Verdict')
    r += 1
    blended_row = r
    ws.cell(row=r, column=1, value='Blended Intrinsic Value')

    r += 1
    upside_row = r
    ws.cell(row=r, column=1, value='Upside / Downside')
    r += 1
    ws.cell(row=r, column=1, value='Confidence Score')
    ws.cell(row=r, column=2, value=result.get('confidence_score', 0)).number_format = '0.00'

    # DCF Scenarios
    r += 2
    _style_section(ws, r, 5, 'DCF Scenarios')
    r += 1
    for label, key in [('Bull Case', 'dcf_bull'), ('Base Case', 'dcf_base'), ('Bear Case', 'dcf_bear')]:
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=result.get(key, 0) or 0).number_format = NUM_FMT
        _fml(ws, r, 3, f'=B{r}/B{cmp_row}-1', PCT_FMT)
        r += 1

    # Method breakdown
    r += 1
    _style_section(ws, r, 5, 'Valuation Methods')
    r += 1
    ws.cell(row=r, column=1, value='Method')
    ws.cell(row=r, column=2, value='Value (Rs)')
    ws.cell(row=r, column=3, value='Weight')
    ws.cell(row=r, column=4, value='Contribution')
    _style_header(ws, r, 4)

    blend = result.get('blend_weights', {'dcf': 0.6, 'relative': 0.3, 'monte_carlo': 0.1})

    r += 1
    method_start = r
    # DCF — reference DCF Model sheet
    ws.cell(row=r, column=1, value='DCF (Base)')
    if refs.get('dcf_iv_row'):
        _fml(ws, r, 2, f"='DCF Model'!B{refs['dcf_iv_row']}", NUM_FMT)
    else:
        ws.cell(row=r, column=2, value=result.get('dcf_base', 0) or 0).number_format = NUM_FMT
    _inp(ws, r, 3, blend.get('dcf', 0.6), PCT_FMT)
    _fml(ws, r, 4, f'=B{r}*C{r}', NUM_FMT)

    r += 1
    # Relative — reference Relative Val sheet
    ws.cell(row=r, column=1, value='Relative')
    if refs.get('relative_adj_row'):
        _fml(ws, r, 2, f"='Relative Val'!B{refs['relative_adj_row']}", NUM_FMT)
    else:
        ws.cell(row=r, column=2, value=result.get('relative_value', 0) or 0).number_format = NUM_FMT
    _inp(ws, r, 3, blend.get('relative', 0.3), PCT_FMT)
    _fml(ws, r, 4, f'=B{r}*C{r}', NUM_FMT)

    r += 1
    # Monte Carlo — reference Monte Carlo sheet
    ws.cell(row=r, column=1, value='Monte Carlo')
    if refs.get('mc_median_row'):
        _fml(ws, r, 2, f"='Monte Carlo'!B{refs['mc_median_row']}", NUM_FMT)
    else:
        ws.cell(row=r, column=2, value=result.get('mc_median', 0) or 0).number_format = NUM_FMT
    _inp(ws, r, 3, blend.get('monte_carlo', 0.1), PCT_FMT)
    _fml(ws, r, 4, f'=B{r}*C{r}', NUM_FMT)
    method_end = r

    # Blended total
    r += 1
    ws.cell(row=r, column=1, value='Blended Total').font = BOLD_FONT
    cell = _fml(ws, r, 4, f'=SUM(D{method_start}:D{method_end})', NUM_FMT)
    cell.font = BOLD_FONT
    blended_formula_row = r

    # Now set the blended value and upside formulas
    _fml(ws, blended_row, 2, f'=D{blended_formula_row}', NUM_FMT)
    ws.cell(row=blended_row, column=2).font = BOLD_FONT
    _fml(ws, upside_row, 2, f'=B{blended_row}/B{cmp_row}-1', PCT_FMT)

    # Driver Hierarchy (4-level) - Summary only, details in Driver Hierarchy tab
    r += 2
    _style_section(ws, r, 5, 'Driver Hierarchy Summary')
    _remark(ws, r, 5, "See 'Driver Hierarchy' tab for details")

    hierarchy = result.get('driver_hierarchy', {})
    outlook = result.get('sector_outlook', {})

    r += 1
    ws.cell(row=r, column=1, value='Outlook')
    ws.cell(row=r, column=2, value=outlook.get('outlook_label', 'NEUTRAL')).font = BOLD_FONT
    ws.cell(row=r, column=3, value=f"Score: {outlook.get('outlook_score', 0):.4f}")

    r += 1
    ws.cell(row=r, column=1, value='Growth Adjustment')
    ws.cell(row=r, column=2, value=outlook.get('growth_adjustment', 0)).number_format = PCT_FMT

    r += 1
    ws.cell(row=r, column=1, value='Margin Adjustment')
    ws.cell(row=r, column=2, value=outlook.get('margin_adjustment', 0)).number_format = PCT_FMT
